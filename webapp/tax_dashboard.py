"""
Tax & P&L dashboard - combined view of equity + MF + SGB + options
+ all transaction costs across all years.

Data sources (mutually exclusive):
  - The user's own Angel One Tax PNL xlsx files in data/tax_pnl/ (default)
  - An ephemeral uploaded session (data/tax_pnl_uploads/<id>/) selected
    via the ?session=<id> query param

Renders as a single pie chart with hover tooltips, plus a year-by-year
breakdown and a P&L summary.

Upload routes in this module accept xlsx (Angel One), csv (Zerodha), and
arbitrary tabular files with a user-supplied column mapping (Generic).
"""
from __future__ import annotations

import json
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse

from webapp import TEMPLATES_DIR
from webapp.data import _portfolio_tickers
from pipeline.tax_pnl import parse_files as parse_uploaded_files
from pipeline.tax_pnl.sessions import (
    MAX_FILES_PER_SESSION,
    SessionMeta,
    cache_parsed,
    create_session,
    delete_session,
    get_cached_parsed,
    get_session,
    list_session_files,
    save_uploaded_file,
    set_column_mapping,
    update_meta,
    validate_upload,
)
from pipeline.tax_pnl.adapters import (
    all_supported_brokers,
    get_adapter,
    get_generic_adapter,
)
from pipeline.tax_pnl.adapters.generic import MAPPING_FIELDS, extract_headers
from pipeline.tax_pnl.report import build_markdown_report

# Where the Tax PNL xlsx files live. Drop new files into this folder
# after downloading from Angel One; the Tax & P&L page picks them up
# on next refresh.
TAX_PNL_DIR = Path(__file__).resolve().parent.parent / "data" / "tax_pnl"
TAX_PNL_PATTERNS = [
    "Tax PNL 2022-23 (1).xlsx",
    "Tax PNL 2023-24 (1).xlsx",
    "Tax PNL 2024-25 (3).xlsx",
    "Tax PNL 2025-26 (1).xlsx",
    "Tax PNL 2026-27.xlsx",
]

router = APIRouter()


def extract_tax_data(force_refresh: bool = False, session_id: str = None) -> dict:
    """Parse Tax PNL xlsx files.

    Args:
        force_refresh: re-parse even if cache is fresh
        session_id:    if set, parse the uploaded session instead of the
                       local files in data/tax_pnl/

    Returns a dict in the original tax_dashboard shape (totals, by_fy,
    all_equity_buys, all_equity_sells), so existing template code
    doesn't need to change.

    Cached at /tmp/full_tax_data.json for local mode, or
    data/tax_pnl_uploads/<session_id>/parsed.json for session mode.
    """
    if session_id:
        return _extract_session_data(session_id, force_refresh)
    return _extract_local_data(force_refresh)


def _extract_local_data(force_refresh: bool) -> dict:
    """The original implementation: parse files from data/tax_pnl/."""
    cache_path = Path("/tmp/full_tax_data.json")
    if not force_refresh and cache_path.exists():
        age = (datetime.now() - datetime.fromtimestamp(cache_path.stat().st_mtime)).total_seconds()
        if age < 240:  # 4 minutes
            try:
                return json.loads(cache_path.read_text())
            except Exception:
                pass

    totals = {
        "equity_buy_value": 0.0, "equity_sell_value": 0.0, "equity_pnl": 0.0,
        "equity_stcg": 0.0, "equity_ltcg": 0.0,
        "equity_stamp_duty": 0.0, "equity_stt": 0.0,
        "equity_brokerage": 0.0, "equity_other_charges": 0.0,
        "equity_intraday_pnl": 0.0,
        "fno_options_turnover": 0.0, "fno_options_pnl": 0.0,
        "fno_futures_turnover": 0.0, "fno_futures_pnl": 0.0,
        "fno_stt": 0.0, "fno_charges": 0.0, "fno_brokerage": 0.0,
        "dividend_income": 0.0,
        "open_holdings_cost": 0.0, "open_holdings_market_value": 0.0,
        "open_holdings_unrealised": 0.0,
        "open_holdings_st_unrealised": 0.0, "open_holdings_lt_unrealised": 0.0,
    }
    fy_data = []
    all_buys, all_sells = [], []

    for fname in TAX_PNL_PATTERNS:
        path = TAX_PNL_DIR / fname
        if not path.exists():
            continue
        fy_label = fname.replace("Tax PNL ", "").replace(".xlsx", "").strip()
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except Exception:
            continue
        fy = {k: 0.0 for k in totals.keys()}
        fy["fy"] = fy_label

        # Equity+Bonds+SGB sheet
        try:
            ws = wb["Equity+Bonds+SGB Trade Details"]
            rows = list(ws.iter_rows(values_only=True))
            section = None
            for row in rows:
                if not row or all(v is None or v == "" for v in row):
                    continue
                first = str(row[0] or "").strip()
                # Summary
                if first == "Net P&L":
                    fy["equity_pnl"] = _num(row[1])
                elif first == "Taxable Delivery P&L (LTCG) Excluding Buyback":
                    fy["equity_ltcg"] = _num(row[1])
                elif first == "Taxable Delivery P&L (STCG) Excluding Buyback":
                    fy["equity_stcg"] = _num(row[1])
                elif first == "Taxable Intraday  P&L (Speculative)":
                    fy["equity_intraday_pnl"] = _num(row[1])
                elif first == "Total Charges and Statutory Levies":
                    fy["equity_other_charges"] = _num(row[1])
                elif first == "Total STT":
                    fy["equity_stt"] = _num(row[1])
                elif first == "Additional Brokerage":
                    fy["equity_brokerage"] = _num(row[1])
                # Section markers
                elif "Intraday" in first:
                    section = "intraday"; continue
                elif "Delivery P&L" in first:
                    section = "delivery"; continue
                elif "Buyback" in first:
                    section = "buyback"; continue
                elif "Transfer" in first:
                    section = "transfer"; continue
                elif "Open Sell" in first:
                    section = "open_sell"; continue
                elif "Open Holdings" in first:
                    section = "open"; continue
                elif first == "ISIN" and len(row) > 5:
                    continue
                isin = row[0] if isinstance(row[0], str) and row[0].startswith("INE") else None
                if not isin: continue
                scrip = row[1] if len(row) > 1 else None
                if not scrip: continue
                if section == "delivery" and len(row) > 12:
                    qty = _num(row[2]); buy_val = _num(row[6])
                    sell_val = _num(row[8]); charges = _num(row[10]) if len(row) > 10 else 0
                    stt = _num(row[11]) if len(row) > 11 else 0
                    if qty > 0:
                        fy["equity_buy_value"] += buy_val
                        fy["equity_sell_value"] += sell_val
                        fy["equity_stamp_duty"] += charges
                        fy["equity_stt"] += stt
                        all_buys.append({"fy": fy_label, "date": str(row[3]) if row[3] else None,
                                         "scrip": scrip, "qty": qty, "buy_val": buy_val})
                        if row[4]:
                            all_sells.append({"fy": fy_label, "date": str(row[4]),
                                              "scrip": scrip, "qty": qty,
                                              "sell_val": sell_val, "pnl": _num(row[12])})
                elif section == "open" and len(row) > 10:
                    qty = _num(row[2]); buy_val = _num(row[4])
                    closing = _num(row[7]) if len(row) > 7 else 0
                    st_un = _num(row[9]) if len(row) > 9 else 0
                    lt_un = _num(row[10]) if len(row) > 10 else 0
                    if qty > 0 and buy_val > 1000:
                        fy["open_holdings_cost"] += buy_val
                        fy["open_holdings_market_value"] += closing * qty
                        fy["open_holdings_st_unrealised"] += st_un
                        fy["open_holdings_lt_unrealised"] += lt_un
        except Exception:
            pass

        # Derivatives
        try:
            ws2 = wb["Derivatives Trade Details"]
            for row in ws2.iter_rows(min_row=2, max_row=10, values_only=True):
                if not row or not row[0]: continue
                first = str(row[0] or "").strip()
                if first == "Futures Turnover": fy["fno_futures_turnover"] = _num(row[1])
                elif first == "Options Turnover": fy["fno_options_turnover"] = _num(row[1])
                elif first == "Taxable Futures P&L (Non Speculative)": fy["fno_futures_pnl"] = _num(row[1])
                elif first == "Taxable Options P&L (Non Speculative)": fy["fno_options_pnl"] = _num(row[1])
                elif first == "Total Charges and Statutory Levies": fy["fno_charges"] = _num(row[1])
                elif first == "Total STT": fy["fno_stt"] = _num(row[1])
        except Exception:
            pass

        # Dividend
        try:
            ws3 = wb["Dividend Report"]
            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True):
                if row and len(row) > 5 and isinstance(row[5], (int, float)):
                    fy["dividend_income"] += row[5]
        except Exception:
            pass

        fy_data.append(fy)
        wb.close()

    for fy in fy_data:
        for k, v in fy.items():
            if k != "fy" and k in totals:
                totals[k] += v

    totals["open_holdings_unrealised"] = (
        totals["open_holdings_st_unrealised"] +
        totals["open_holdings_lt_unrealised"]
    )

    out = {"totals": totals, "by_fy": fy_data,
           "all_equity_buys": all_buys, "all_equity_sells": all_sells}
    try:
        cache_path.write_text(json.dumps(out, indent=2, default=str))
    except Exception:
        pass
    return out


def _extract_session_data(session_id: str, force_refresh: bool) -> dict:
    """Extract tax data for an ephemeral uploaded session.

    - Validates the session exists and isn't expired
    - Uses the cached parse if fresh (10 min TTL)
    - Otherwise re-runs all adapters against the uploaded files
    - Returns the SAME dict shape as _extract_local_data so the
      tax_page() template doesn't need to know which mode it's in
    """
    meta = get_session(session_id)
    if meta is None:
        raise HTTPException(404, f"session not found or expired: {session_id}")

    if not force_refresh:
        cached = get_cached_parsed(session_id)
        if cached is not None:
            return cached

    files = list_session_files(session_id)
    if not files:
        out = {
            "totals": _empty_totals(),
            "by_fy": [],
            "all_equity_buys": [],
            "all_equity_sells": [],
            "_session": _session_summary(meta, []),
        }
        cache_parsed(session_id, out)
        return out

    column_mapping = meta.column_mapping or {}
    if column_mapping:
        from pipeline.tax_pnl import parse_files_with_mapping
        data = parse_files_with_mapping(files, column_mapping, label=meta.label)
    else:
        data = parse_uploaded_files(files, label=meta.label)

    update_meta(session_id,
                source_files=data.source_files,
                detected_brokers=data.detected_brokers)

    totals = data.totals()
    by_fy = data.by_fy()
    all_sells = []
    for t in data.trades:
        avg_buy = t.buy_value / t.quantity if t.quantity else 0
        avg_sell = t.sell_value / t.quantity if t.quantity else 0
        all_sells.append({
            "fy": t.fy,
            "date": t.sell_date.isoformat() if t.sell_date else None,
            "buy_date": t.buy_date.isoformat() if t.buy_date else "—",
            "scrip": t.scrip,
            "qty": t.quantity,
            "sell_val": t.sell_value,
            "pnl": t.pnl,
            "avg_buy": avg_buy,
            "avg_sell": avg_sell,
            "source_broker": t.source_broker,
        })
    all_buys = [{
        "fy": t.fy, "date": t.buy_date.isoformat() if t.buy_date else None,
        "scrip": t.scrip, "qty": t.quantity, "buy_val": t.buy_value,
        "source_broker": t.source_broker,
    } for t in data.trades if t.buy_date]

    out = {
        "totals": totals,
        "by_fy": by_fy,
        "all_equity_buys": all_buys,
        "all_equity_sells": all_sells,
        "_session": _session_summary(meta, data),
    }
    cache_parsed(session_id, out)
    return out


def _empty_totals() -> dict:
    return {
        "equity_buy_value": 0.0, "equity_sell_value": 0.0, "equity_pnl": 0.0,
        "equity_stcg": 0.0, "equity_ltcg": 0.0,
        "equity_stamp_duty": 0.0, "equity_stt": 0.0,
        "equity_brokerage": 0.0, "equity_other_charges": 0.0,
        "equity_intraday_pnl": 0.0,
        "fno_options_turnover": 0.0, "fno_options_pnl": 0.0,
        "fno_futures_turnover": 0.0, "fno_futures_pnl": 0.0,
        "fno_stt": 0.0, "fno_charges": 0.0, "fno_brokerage": 0.0,
        "dividend_income": 0.0,
        "open_holdings_cost": 0.0, "open_holdings_market_value": 0.0,
        "open_holdings_unrealised": 0.0,
        "open_holdings_st_unrealised": 0.0, "open_holdings_lt_unrealised": 0.0,
    }


def _session_summary(meta: SessionMeta, data) -> dict:
    return {
        "session_id": meta.session_id,
        "label": meta.label,
        "created_at": meta.created_at,
        "expires_at": meta.expires_at,
        "source_files": getattr(data, "source_files", meta.source_files),
        "detected_brokers": getattr(data, "detected_brokers", meta.detected_brokers),
        "parse_warnings": getattr(data, "parse_warnings", []),
        "is_uploaded": True,
    }


def _num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


# ---------- Routes ----------
@router.get("/api/tax")
def api_tax() -> dict:
    """JSON: all tax data, refreshable."""
    return extract_tax_data(force_refresh=True)


@router.get("/api/tax/cached")
def api_tax_cached() -> dict:
    """JSON: cached tax data (4-min TTL)."""
    return extract_tax_data(force_refresh=False)


@router.get("/api/tax/trades")
def api_tax_trades() -> dict:
    """JSON: all closed equity trades, enriched with current LTP, verdict.
    Returns BOTH per-trade detail AND per-stock aggregates.
    """
    data = extract_tax_data(force_refresh=False)
    sells = data["all_equity_sells"]

    ltp_path = Path("/tmp/current_ltps.json")
    if ltp_path.exists():
        try:
            current_ltps = json.loads(ltp_path.read_text())
        except Exception:
            current_ltps = {}
    else:
        current_ltps = {}

    def verdict_for(pnl_pct: float) -> tuple[str, str, str]:
        if pnl_pct > 20:   return ("GREAT SELL",   "#2f7a3d", "is-positive")
        if pnl_pct > 5:    return ("GOOD SELL",    "#5cb85c", "is-positive")
        if pnl_pct > 0:    return ("OK SELL",      "#888",    "text-muted")
        if pnl_pct > -5:   return ("BAD TIMING",   "#d58512", "is-negative")
        if pnl_pct > -20:  return ("POOR SELL",    "#b3382c", "is-negative")
        return ("TERRIBLE SELL", "#7a1313", "is-negative")

    def enrich(t: dict) -> dict:
        scrip = t.get("scrip", "")
        ltp_info = current_ltps.get(scrip, {}) or {}
        cur_ltp = ltp_info.get("ltp")
        avg_sell = t.get("avg_sell", 0) or 0
        avg_buy = t.get("avg_buy", 0) or 0
        qty = t.get("qty", 0) or 0
        pnl = t.get("pnl", 0) or 0
        sell_val = t.get("sell_val", 0) or 0
        pnl_pct = (pnl / sell_val * 100) if sell_val else 0
        v_label, v_color, v_class = verdict_for(pnl_pct)
        if cur_ltp and avg_sell:
            hyp_pnl = (cur_ltp - avg_sell) * qty
            hyp_pct = ((cur_ltp - avg_sell) / avg_sell) * 100 if avg_sell else 0
        else:
            hyp_pnl = None
            hyp_pct = None
        return {**t,
                "cur_ltp": cur_ltp,
                "52w_high": ltp_info.get("52w_high"),
                "52w_low": ltp_info.get("52w_low"),
                "exchange": ltp_info.get("exchange"),
                "pnl_pct": pnl_pct,
                "verdict": v_label, "verdict_color": v_color, "verdict_class": v_class,
                "hypothetical_pnl": hyp_pnl, "hypothetical_pct": hyp_pct}

    enriched = [enrich(t) for t in sells]
    enriched.sort(key=lambda t: t.get("pnl_pct", 0))

    # ---------- Per-stock aggregation ----------
    # Group trades by scrip. For each stock, compute:
    #   total_qty_sold, weighted_avg_buy, weighted_avg_sell, total_pnl,
    #   total_sell_value, total_buy_value, first_buy_date, last_sell_date,
    #   current_ltp, weighted_avg_held_pnl, verdict (based on overall pnl%)
    by_stock: dict[str, list[dict]] = {}
    for t in enriched:
        by_stock.setdefault(t["scrip"], []).append(t)

    aggregated = []
    for scrip, trades in by_stock.items():
        total_qty = sum(t["qty"] for t in trades)
        total_buy_val = sum((t.get("avg_buy", 0) or 0) * t["qty"] for t in trades)
        total_sell_val = sum(t.get("sell_val", 0) or 0 for t in trades)
        total_pnl = sum(t.get("pnl", 0) or 0 for t in trades)
        wavg_buy = total_buy_val / total_qty if total_qty else 0
        wavg_sell = total_sell_val / total_qty if total_qty else 0
        # Dates
        buy_dates = sorted([t["buy_date"] for t in trades
                            if t.get("buy_date") and t["buy_date"] != "—"])
        sell_dates = sorted([t["date"] for t in trades
                             if t.get("date") and t["date"] != "—"])
        first_buy = buy_dates[0] if buy_dates else "—"
        last_sell = sell_dates[-1] if sell_dates else "—"
        # Current LTP and hypothetical
        ltp_info = current_ltps.get(scrip, {}) or {}
        cur_ltp = ltp_info.get("ltp")
        hyp_pnl = ((cur_ltp - wavg_sell) * total_qty) if (cur_ltp and wavg_sell) else None
        hyp_pct = (((cur_ltp - wavg_sell) / wavg_sell) * 100
                    if (cur_ltp and wavg_sell) else None)
        pnl_pct = (total_pnl / total_sell_val * 100) if total_sell_val else 0
        v_label, v_color, v_class = verdict_for(pnl_pct)
        # Renamed stock note
        note = ""
        if scrip == "TATAMOTORS":
            note = "TATAMOTORS DVR demerged in 2024-25. Current: Tata Motors Passenger Vehicles (NSE: TMPV)."
        elif scrip == "ZOMATO":
            note = "Zomato Ltd rebranded to Eternal Ltd in Mar 2025 (NSE: ETERNAL)."
        aggregated.append({
            "scrip": scrip,
            "trade_count": len(trades),
            "first_buy_date": first_buy,
            "last_sell_date": last_sell,
            "wavg_buy": wavg_buy,
            "wavg_sell": wavg_sell,
            "total_qty": total_qty,
            "total_buy_value": total_buy_val,
            "total_sell_value": total_sell_val,
            "total_pnl": total_pnl,
            "pnl_pct": pnl_pct,
            "verdict": v_label,
            "verdict_color": v_color,
            "verdict_class": v_class,
            "cur_ltp": cur_ltp,
            "hypothetical_pnl": hyp_pnl,
            "hypothetical_pct": hyp_pct,
            "note": note,
        })
    # Sort aggregated by pnl ascending (worst first)
    aggregated.sort(key=lambda a: a["pnl_pct"])

    # ---------- Summary stats ----------
    verdicts = {}
    for t in enriched:
        verdicts[t["verdict"]] = verdicts.get(t["verdict"], 0) + 1
    total_pnl = sum(t.get("pnl", 0) or 0 for t in enriched)
    best = max(enriched, key=lambda t: t.get("pnl_pct", 0)) if enriched else None
    worst = min(enriched, key=lambda t: t.get("pnl_pct", 0)) if enriched else None

    return {
        "trades": enriched,                      # 57 individual rows
        "by_stock": aggregated,                 # 28 unique stocks
        "summary": {
            "total_trades": len(enriched),
            "unique_stocks": len(aggregated),
            "total_pnl": total_pnl,
            "verdicts": verdicts,
            "best": {
                "scrip": best["scrip"], "pnl": best["pnl"],
                "pnl_pct": best["pnl_pct"], "date": best["date"],
            } if best else None,
            "worst": {
                "scrip": worst["scrip"], "pnl": worst["pnl"],
                "pnl_pct": worst["pnl_pct"], "date": worst["date"],
            } if worst else None,
            "ltps_available": sum(1 for t in enriched if t.get("cur_ltp")),
        },
        "asof": datetime.now().isoformat(timespec="seconds"),
    }


@router.get("/tax", response_class=HTMLResponse)
def tax_page(request: Request, session: str = None):
    """Render the comprehensive tax & P&L dashboard.

    Query params:
        session=<id>  If set, analyze an uploaded ephemeral session
                      instead of the local files in data/tax_pnl/.
    """
    from webapp.server import templates

    data = extract_tax_data(force_refresh=False, session_id=session)
    totals = data["totals"]
    fy_data = data["by_fy"]
    buys = data["all_equity_buys"]
    sells = data["all_equity_sells"]
    session_info = data.get("_session")

    # ============ Compute the big pie chart data ============
    # Pie chart slices (we want the user to see ALL major money flows):
    pie = []

    # 1. Equity buy value (money out)
    if totals["equity_buy_value"] > 0:
        pie.append({
            "label": "Equity (bought & sold)",
            "value": abs(totals["equity_buy_value"]),
            "color": "#4e79a7",
            "detail": f"₹{totals['equity_buy_value']:,.0f} bought across 4 years",
        })

    # 2. Equity sell value (money in, gross)
    if totals["equity_sell_value"] > 0:
        pie.append({
            "label": "Equity (sell proceeds)",
            "value": abs(totals["equity_sell_value"]),
            "color": "#59a14f",
            "detail": f"₹{totals['equity_sell_value']:,.0f} received from sales",
        })

    # 3. Realised equity P&L
    if totals["equity_pnl"] != 0:
        pie.append({
            "label": "Realised P&L (STCG+LTCG)",
            "value": abs(totals["equity_pnl"]),
            "color": "#edc948" if totals["equity_pnl"] > 0 else "#e15759",
            "detail": (f"+₹{totals['equity_pnl']:,.0f} profit" if totals["equity_pnl"] > 0
                       else f"-₹{abs(totals['equity_pnl']):,.0f} loss"),
        })

    # 4. F&O / options turnover
    if totals["fno_options_turnover"] > 0:
        pie.append({
            "label": "Options turnover",
            "value": abs(totals["fno_options_turnover"]),
            "color": "#b07aa1",
            "detail": f"₹{totals['fno_options_turnover']:,.0f} traded (premium paid)",
        })

    # 5. Options P&L
    if totals["fno_options_pnl"] != 0:
        pie.append({
            "label": "Options P&L",
            "value": abs(totals["fno_options_pnl"]),
            "color": "#76b7b2" if totals["fno_options_pnl"] > 0 else "#ff9da7",
            "detail": (f"+₹{totals['fno_options_pnl']:,.0f}" if totals['fno_options_pnl'] > 0
                       else f"-₹{abs(totals['fno_options_pnl']):,.0f} loss on options"),
        })

    # 6. Open holdings cost
    if totals["open_holdings_cost"] > 0:
        pie.append({
            "label": "Current holdings (cost)",
            "value": abs(totals["open_holdings_cost"]),
            "color": "#ff9da7",
            "detail": f"₹{totals['open_holdings_cost']:,.0f} invested, still held",
        })

    # 7. Unrealised loss
    if totals["open_holdings_unrealised"] != 0:
        pie.append({
            "label": "Unrealised P&L (open)",
            "value": abs(totals["open_holdings_unrealised"]),
            "color": "#9c755f" if totals["open_holdings_unrealised"] > 0 else "#bab0ac",
            "detail": (f"+₹{totals['open_holdings_unrealised']:,.0f} unrealised profit"
                       if totals['open_holdings_unrealised'] > 0
                       else f"-₹{abs(totals['open_holdings_unrealised']):,.0f} unrealised loss"),
        })

    # 8. Dividends received
    if totals["dividend_income"] > 0:
        pie.append({
            "label": "Dividend income",
            "value": totals["dividend_income"],
            "color": "#2ca02c",
            "detail": f"₹{totals['dividend_income']:,.0f} in dividends",
        })

    # 9. STT paid
    if totals["equity_stt"] > 0:
        pie.append({
            "label": "STT (equity)",
            "value": totals["equity_stt"],
            "color": "#d62728",
            "detail": f"₹{totals['equity_stt']:,.0f} STT paid",
        })

    # 10. F&O STT
    if totals["fno_stt"] > 0:
        pie.append({
            "label": "STT (F&O)",
            "value": totals["fno_stt"],
            "color": "#ff7f0e",
            "detail": f"₹{totals['fno_stt']:,.0f} F&O STT",
        })

    # 11. Stamp duty
    if totals["equity_stamp_duty"] > 0:
        pie.append({
            "label": "Stamp duty",
            "value": totals["equity_stamp_duty"],
            "color": "#9467bd",
            "detail": f"₹{totals['equity_stamp_duty']:,.0f} stamp duty",
        })

    # 12. Other charges (GST, SEBI, etc.)
    if totals["equity_other_charges"] > 0:
        pie.append({
            "label": "Other equity charges",
            "value": totals["equity_other_charges"],
            "color": "#8c564b",
            "detail": "GST, SEBI fees, AMC, etc.",
        })

    # 13. F&O charges
    if totals["fno_charges"] > 0:
        pie.append({
            "label": "F&O charges",
            "value": totals["fno_charges"],
            "color": "#e377c2",
            "detail": f"₹{totals['fno_charges']:,.0f} F&O charges + brokerage",
        })

    # ============ Compute totals ============
    # Total money out
    total_money_out = (
        totals["equity_buy_value"] +
        totals["fno_options_turnover"] +
        totals["equity_stt"] +
        totals["equity_stamp_duty"] +
        totals["equity_other_charges"] +
        totals["fno_stt"] +
        totals["fno_charges"] +
        totals["open_holdings_cost"]
    )

    # Total money back
    total_money_back = (
        totals["equity_sell_value"] +
        totals["dividend_income"]
    )

    # Net realised P&L (closed trades only)
    net_realised = (
        totals["equity_pnl"] +
        totals["fno_options_pnl"] +
        totals["fno_futures_pnl"] +
        totals["equity_intraday_pnl"]
    )

    # Net unrealised (open positions)
    net_unrealised = totals["open_holdings_unrealised"]

    # Total profit/loss (realised + unrealised + dividend)
    total_gain_loss = net_realised + net_unrealised + totals["dividend_income"]

    # Total transaction costs
    total_costs = (
        totals["equity_stt"] + totals["fno_stt"] +
        totals["equity_stamp_duty"] + totals["equity_other_charges"] +
        totals["fno_charges"] + totals["fno_brokerage"]
    )

    return templates.TemplateResponse(request, "tax.html", _ctx_for_tax(
        active_nav="tax",
        page_title="Tax & P&L Dashboard",
        snapshot={
            "asof": datetime.now().isoformat(timespec="seconds"),
            "asof_human": "just now",
        },
        pie=pie,
        totals=totals,
        fy_data=fy_data,
        buys=buys,
        sells=sells,
        total_money_out=total_money_out,
        total_money_back=total_money_back,
        net_realised=net_realised,
        net_unrealised=net_unrealised,
        total_gain_loss=total_gain_loss,
        total_costs=total_costs,
        session=session_info,
    ))


def _ctx_for_tax(**extra) -> dict:
    """Build template context like webapp.server._ctx()."""
    return {
        "active_nav": None,
        "page_title": "Portfolio Tracker",
        **extra,
    }


# ============================================================
# Ephemeral upload routes
# ============================================================
# These let the user upload one or more Tax P&L files (Angel One xlsx,
# Zerodha CSV, or any tabular file with manual column mapping) and get
# the same Tax & P&L dashboard rendered against the upload. Sessions
# live 24h and live in data/tax_pnl_uploads/<id>/ — outside the
# user's own data/tax_pnl/ so they can never pollute the existing
# pipeline.

@router.get("/api/tax/upload/brokers")
def upload_brokers() -> dict:
    """List of supported brokers for the upload UI."""
    return {"brokers": all_supported_brokers()}


@router.get("/api/tax/upload/headers")
def upload_headers(session: str) -> dict:
    """Return the column headers of an uploaded file. Used by the
    column-mapping UI for the Generic adapter.

    The session must already exist and have at least one file uploaded.
    """
    meta = get_session(session)
    if meta is None:
        raise HTTPException(404, "session not found or expired")
    files = list_session_files(session)
    if not files:
        raise HTTPException(400, "no files in session")
    try:
        headers = extract_headers(files[0])
    except Exception as e:
        raise HTTPException(400, f"could not read headers: {e}")
    return {"headers": headers, "fields": MAPPING_FIELDS,
            "file": files[0].name}


@router.post("/api/tax/upload")
async def upload_tax_pnl(
    request: Request,
    label: str = Form(None),
    column_mapping: str = Form(None),
    files: list[UploadFile] = File(...),
):
    """Upload one or more Tax P&L files. Creates a new ephemeral session."""
    if not files:
        raise HTTPException(400, "no files uploaded")
    if len(files) > MAX_FILES_PER_SESSION:
        raise HTTPException(400,
            f"too many files: {len(files)} (max {MAX_FILES_PER_SESSION})")

    mapping = {}
    if column_mapping:
        try:
            mapping = json.loads(column_mapping)
            if not isinstance(mapping, dict):
                raise ValueError("column_mapping must be a JSON object")
        except (json.JSONDecodeError, ValueError) as e:
            raise HTTPException(400, f"invalid column_mapping: {e}")

    meta = create_session(label=label)
    if mapping:
        set_column_mapping(meta.session_id, mapping)

    files_uploaded = []
    files_rejected = []

    for upload in files:
        try:
            content = await upload.read()
        except Exception as e:
            files_rejected.append({"name": upload.filename or "?", "error": str(e)})
            continue
        err = validate_upload(upload.filename or "", content)
        if err is not None:
            files_rejected.append({"name": upload.filename or "?", "error": err})
            continue
        with tempfile.NamedTemporaryFile(
            delete=False, suffix=Path(upload.filename).suffix.lower()
        ) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)
        try:
            dest = save_uploaded_file(meta.session_id, tmp_path, upload.filename)
            files_uploaded.append(dest.name)
        except Exception as e:
            files_rejected.append({"name": upload.filename or "?", "error": str(e)})
            try:
                tmp_path.unlink()
            except Exception:
                pass

    if not files_uploaded and not mapping:
        delete_session(meta.session_id)
        raise HTTPException(400, {
            "error": "no files accepted",
            "rejected": files_rejected,
        })

    return {
        "session_id": meta.session_id,
        "label": meta.label,
        "created_at": meta.created_at,
        "expires_at": meta.expires_at,
        "ttl_hours": 24,
        "files_uploaded": files_uploaded,
        "files_rejected": files_rejected,
        "tax_url": f"/tax?session={meta.session_id}",
    }


@router.post("/api/tax/upload/{session_id}/mapping")
async def set_mapping(session_id: str, mapping: str = Form(...)):
    """Set or update the column mapping for a Generic-adapter session."""
    meta = get_session(session_id)
    if meta is None:
        raise HTTPException(404, "session not found or expired")
    try:
        new_mapping = json.loads(mapping)
        if not isinstance(new_mapping, dict):
            raise ValueError("mapping must be a JSON object")
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(400, f"invalid mapping: {e}")
    set_column_mapping(session_id, new_mapping)
    extract_tax_data(force_refresh=True, session_id=session_id)
    return {"ok": True, "session_id": session_id}


@router.delete("/api/tax/upload/{session_id}")
def delete_upload_session(session_id: str):
    """Delete an ephemeral session and all its files."""
    deleted = delete_session(session_id)
    if not deleted:
        raise HTTPException(404, "session not found or already deleted")
    return {"ok": True, "session_id": session_id}


@router.get("/api/tax/upload/{session_id}/report", response_class=PlainTextResponse)
def upload_report(session_id: str) -> PlainTextResponse:
    """Return the markdown report for an uploaded session."""
    meta = get_session(session_id)
    if meta is None:
        raise HTTPException(404, "session not found or expired")
    files = list_session_files(session_id)
    if not files:
        raise HTTPException(400, "no files in session")
    column_mapping = meta.column_mapping or {}
    if column_mapping:
        from pipeline.tax_pnl import parse_files_with_mapping
        data = parse_files_with_mapping(files, column_mapping, label=meta.label)
    else:
        data = parse_uploaded_files(files, label=meta.label)
    md = build_markdown_report(data, label=meta.label)
    return PlainTextResponse(md, media_type="text/markdown")