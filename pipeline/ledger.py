"""
pipeline.ledger
---------------
Build a complete buy/sell ledger for the user's equity positions, walking
across all 5 xlsx files in data/tax_pnl/.

Each ledger entry is a single buy or sell transaction (one row of the
Delivery P&L section). We keep both:
- **Current positions**: open lots that are still held (sell_date is None
  or sell happens in a different xlsx than the buy)
- **Sold positions**: closed lots with both buy and sell dates

The ledger is the raw data the cohort module uses for total return
calculations.
"""
from __future__ import annotations

import glob
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

from pipeline.portfolio_truth import load_truth
from pipeline.marketcap import classify, get_market_cap_cr

PROJECT = Path(__file__).resolve().parents[1]
TAX_DIR = PROJECT / "data" / "tax_pnl"


@dataclass
class LedgerEntry:
    """A single buy or sell event from the tax P&L xlsx."""
    ticker: str
    isin: str = ""
    buy_date: Optional[date] = None
    sell_date: Optional[date] = None
    qty: int = 0
    buy_price: float = 0.0
    sell_price: float = 0.0
    charges: float = 0.0
    buy_value: float = 0.0          # qty * buy_price
    sell_value: float = 0.0         # qty * sell_price
    pnl: float = 0.0               # (sell_price - buy_price) * qty - charges
    holding_days: int = 0
    xlsx: str = ""
    # Cohort classification (filled in later by cohort module)
    cohort: str = ""
    tier: str = ""                # large / mid / small / unknown
    is_etf: bool = False

    @property
    def is_open(self) -> bool:
        return self.sell_date is None

    @property
    def is_closed(self) -> bool:
        return self.sell_date is not None


def _parse_dmy(s: str) -> Optional[date]:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# Known ETFs that should be excluded from the equity cohort comparison
KNOWN_ETFS = {
    "GOLDBEES", "METALIETF", "NEXT50IETF", "AUTOIETF", "PSUBNKBEES",
    "ITBEES", "BANKBEES", "PHARMABEES", "LIQUIDBEES", "NIFTYBEES",
    "CPSEETF", "MAFANG", "MAFSMALLCAP", "SILVERBEES", "GOLD1", "SILVER1",
    "JUNIORBEES", "MIDCAP",  # NSE listed Midcap ETF (don't confuse with midcap stocks)
}


def _is_etf(ticker: str) -> bool:
    t = ticker.upper()
    if t in KNOWN_ETFS:
        return True
    # Heuristic: any ticker ending in ETF, BEES, or matching gold/silver
    if t.endswith("ETF") or t.endswith("BEES"):
        return True
    if t.startswith("GOLD") or t.startswith("SILVER"):
        return True
    return False


def _read_delivery_pnl_sells(ws, xlsx_name: str) -> list[dict]:
    """Read all closed Delivery P&L rows (Buy Date + Sell Date present)."""
    in_delivery = False
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c else "" for c in row]
        # Detect Delivery P&L header
        if (len(cells) > 9 and cells[1] == "Scrip Name"
                and "Buy Date" in cells[3]
                and "Cost Of Acquisition" in cells[9]):
            in_delivery = True
            continue
        if not in_delivery:
            continue
        # Section break
        if cells and cells[0] and any(s in cells[0] for s in (
                "Open Sell", "Open Holdings", "Disclaimer", "Buyback",
                "Transfer", "Intraday (Speculation)", "Calculations")):
            in_delivery = False
            continue
        if not cells[1] or cells[1] == "Sub total":
            continue
        # Parse row
        bd = _parse_dmy(cells[3])
        sd = _parse_dmy(cells[4])
        if bd is None or sd is None:
            continue  # need both buy and sell dates for a closed lot
        try:
            qty = int(cells[2])
            bp = float(cells[5])
            sp = float(cells[7])
            charges = float(cells[10]) if len(cells) > 10 and cells[10] else 0.0
        except (ValueError, TypeError, IndexError):
            continue
        rows.append({
            "isin": cells[0],
            "ticker": cells[1].strip().upper(),
            "buy_date": bd,
            "sell_date": sd,
            "qty": qty,
            "buy_price": bp,
            "sell_price": sp,
            "charges": charges,
        })
    return rows


def _read_invested_lots(ws, xlsx_name: str) -> list[dict]:
    """Read all Delivery P&L rows with Buy Date but no Sell Date.
    These are buys that may not have been sold yet within this xlsx's FY.
    Note: most current positions are NOT in Delivery P&L; they show up
    only in the Open Holdings section. So this is a small subset."""
    in_delivery = False
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c else "" for c in row]
        if (len(cells) > 9 and cells[1] == "Scrip Name"
                and "Buy Date" in cells[3]
                and "Cost Of Acquisition" in cells[9]):
            in_delivery = True
            continue
        if not in_delivery:
            continue
        if cells and cells[0] and any(s in cells[0] for s in (
                "Open Sell", "Open Holdings", "Disclaimer", "Buyback",
                "Transfer", "Intraday (Speculation)", "Calculations")):
            in_delivery = False
            continue
        if not cells[1] or cells[1] == "Sub total":
            continue
        bd = _parse_dmy(cells[3])
        sd = _parse_dmy(cells[4])
        if bd is None or sd is not None:
            continue
        try:
            qty = int(cells[2])
            bp = float(cells[5])
        except (ValueError, TypeError):
            continue
        rows.append({
            "isin": cells[0],
            "ticker": cells[1].strip().upper(),
            "buy_date": bd,
            "sell_date": None,
            "qty": qty,
            "buy_price": bp,
        })
    return rows


def build_ledger() -> list[LedgerEntry]:
    """Walk all 5 xlsx files and build the full buy/sell ledger.

    Returns both closed positions (Delivery P&L with both buy and sell)
    and open positions (from the current truth.json + reconstructed via
    snapshot diffs in cagr.py — but for the ledger we use the simpler
    truth.json + Open Holdings as the source of truth for current
    positions).
    """
    files = sorted(glob.glob(str(TAX_DIR / "Tax PNL *.xlsx")))
    out: list[LedgerEntry] = []

    # 1) Walk all xlsx Delivery P&L for closed positions
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        xlsx_name = Path(f).name
        for r in _read_delivery_pnl_sells(ws, xlsx_name):
            holding_days = (r["sell_date"] - r["buy_date"]).days
            buy_value = r["qty"] * r["buy_price"]
            sell_value = r["qty"] * r["sell_price"]
            pnl = sell_value - buy_value - r["charges"]
            out.append(LedgerEntry(
                ticker=r["ticker"],
                isin=r["isin"],
                buy_date=r["buy_date"],
                sell_date=r["sell_date"],
                qty=r["qty"],
                buy_price=r["buy_price"],
                sell_price=r["sell_price"],
                charges=r["charges"],
                buy_value=buy_value,
                sell_value=sell_value,
                pnl=pnl,
                holding_days=holding_days,
                xlsx=xlsx_name,
                is_etf=_is_etf(r["ticker"]),
            ))

    # 2) Current positions: from truth.json
    # Each ticker has one entry per the current Open Holdings. We treat
    # the buy date as the first-buy-date we can infer (from the earliest
    # xlsx in which the ticker appeared in Open Holdings, clamped to today).
    truth = load_truth()
    # Find the earliest Open Holdings date for each ticker
    earliest_oh: dict[str, date] = {}
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        asof = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if cells and "Open Holdings as of" in cells[0]:
                tail = cells[0].split("as of")[-1].strip()
                d = _parse_dmy(tail)
                if d and d <= date.today():
                    asof = d
                break
        if asof is None:
            continue
        # Now read the open holdings
        in_h = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if not cells:
                continue
            if "Open Holdings" in cells[0]:
                in_h = True
                continue
            if in_h:
                if not cells[1] or cells[1] in ("Scrip Name", "Sub total", ""):
                    if cells[0] and any(x in cells[0] for x in ("Breakup", "Disclaimer", "Calculations")):
                        break
                    continue
                try:
                    qty = int(cells[2])
                except (ValueError, TypeError):
                    continue
                if qty > 0:
                    tk = cells[1].strip().upper()
                    if tk not in earliest_oh or asof < earliest_oh[tk]:
                        earliest_oh[tk] = asof

    for tk, pos in truth.get("equity", {}).items():
        tk = tk.upper()
        qty = pos.get("qty", 0)
        avg = pos.get("avg_price", 0.0)
        if qty <= 0 or avg <= 0:
            continue
        # Use the earliest Open Holdings date as the buy date proxy,
        # clamped to today.
        first_seen = earliest_oh.get(tk)
        if first_seen is None or first_seen > date.today():
            first_seen = date.today()
        out.append(LedgerEntry(
            ticker=tk,
            buy_date=first_seen,
            sell_date=None,
            qty=qty,
            buy_price=avg,
            charges=0.0,
            buy_value=qty * avg,
            xlsx="current_open_holdings",
            is_etf=_is_etf(tk),
        ))

    # 3) Classify cohort and tier for each entry
    for entry in out:
        if entry.is_etf:
            entry.tier = "etf"
            entry.cohort = "etf"
        else:
            entry.tier = classify(entry.ticker)
            # Cohort assignment is done by the cohort module, but for
            # the ledger we tag "equity_<tier>" so the cohort module
            # can group easily.
            entry.cohort = f"equity_{entry.tier}" if entry.tier != "unknown" else "equity_unknown"

    return out


def ledger_summary(ledger: list[LedgerEntry]) -> dict:
    """Group the ledger by cohort and tier. Returns a nested dict."""
    summary: dict = {}
    for e in ledger:
        cohort = e.cohort or "unknown"
        tier = e.tier or "unknown"
        key = (cohort, tier, e.ticker)
        if key not in summary:
            summary[key] = {
                "cohort": cohort,
                "tier": tier,
                "ticker": e.ticker,
                "is_etf": e.is_etf,
                "open_qty": 0,
                "open_value": 0.0,
                "closed_qty": 0,
                "closed_buy_value": 0.0,
                "closed_sell_value": 0.0,
                "closed_pnl": 0.0,
                "lots": 0,
            }
        s = summary[key]
        s["lots"] += 1
        if e.is_open:
            s["open_qty"] += e.qty
            s["open_value"] += e.buy_value
        else:
            s["closed_qty"] += e.qty
            s["closed_buy_value"] += e.buy_value
            s["closed_sell_value"] += e.sell_value
            s["closed_pnl"] += e.pnl
    return summary
