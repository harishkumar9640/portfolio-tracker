"""Generic Tax P&L adapter for arbitrary tabular exports.

This adapter is invoked when neither Angel One nor Zerodha is auto-detected.
The user provides a column mapping in the upload UI:

    {
        "scrip":      "Stock Name",
        "isin":       "ISIN",                  # optional
        "buy_date":   "Purchase Date",
        "sell_date":  "Sale Date",
        "quantity":   "Quantity",
        "buy_value":  "Buy Value",             # total, not per-share
        "sell_value": "Sell Value",            # total, not per-share
        "pnl":        "Realised P&L",          # optional; computed if missing
        "charges":    "Charges",               # optional
        "fy":         "FY",                    # optional; otherwise from filename
    }

The mapping keys are canonical. The values are the literal column names
(or 0-based column indices as strings like "3") in the uploaded file.

For xlsx files: rows are taken from the first non-empty sheet, starting
at the first row that contains the mapped header.
For csv files: same logic, csv.DictReader.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

import openpyxl

from pipeline.tax_pnl import (
    Trade, _num, detect_fy_from_filename, parse_date,
)


def _resolve_index(value: str, headers: list[str]) -> int:
    """Resolve a mapping value (column name or numeric index) to a 0-based index."""
    if value is None:
        return -1
    s = str(value).strip()
    if s.isdigit():
        return int(s)
    norm_headers = {h.strip().lower(): i for i, h in enumerate(headers)}
    return norm_headers.get(s.lower(), -1)


def _read_rows(file: Path) -> tuple[list[str], list[list[Any]]]:
    """Read (headers, data_rows) from a csv or xlsx file. Returns the first
    sheet for xlsx. Skips empty leading lines."""
    if file.suffix.lower() == ".csv":
        with open(file, "r", encoding="utf-8-sig", errors="replace") as f:
            text_lines = []
            for _ in range(10):
                line = f.readline()
                if not line:
                    break
                text_lines.append(line)
            # Find first non-empty line as the header
            start = 0
            for i, l in enumerate(text_lines):
                if l.strip():
                    start = i
                    break
            reader = csv.reader(text_lines[start:])
            rows = list(reader)
            if not rows:
                return [], []
            headers = rows[0]
            data = [r for r in rows[1:] if any(c.strip() for c in r if c)]
            return headers, data
    else:
        wb = openpyxl.load_workbook(str(file), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            headers = None
            data = []
            for row in rows_iter:
                if row is None:
                    continue
                row_list = list(row)
                if headers is None:
                    if any(c is not None and str(c).strip() for c in row_list):
                        headers = [str(c) if c is not None else "" for c in row_list]
                    continue
                if any(c is not None and str(c).strip() for c in row_list):
                    data.append(row_list)
            return headers or [], data
        finally:
            wb.close()


class GenericAdapter:
    name = "generic"

    def __init__(self, column_mapping: dict[str, str]):
        self.column_mapping = column_mapping

    def can_parse(self, file: Path) -> bool:
        # The generic adapter is opt-in only — auto-detect never returns it.
        return False

    def parse(self, file: Path) -> dict:
        fy_label = (self.column_mapping.get("fy")
                    or detect_fy_from_filename(file.name))
        headers, rows = _read_rows(file)
        if not headers or not rows:
            return {
                "fy_summaries": {fy_label: _empty_fy_dict()},
                "trades": [],
                "open_holdings": [],
            }

        # Resolve column indices
        idx: dict[str, int] = {}
        for canon, mapping_value in self.column_mapping.items():
            if canon in ("fy",) or mapping_value in (None, ""):
                continue
            idx[canon] = _resolve_index(mapping_value, headers)

        required = ("scrip", "quantity", "buy_value", "sell_value")
        if not all(idx.get(r, -1) >= 0 for r in required):
            return {
                "fy_summaries": {fy_label: _empty_fy_dict()},
                "trades": [],
                "open_holdings": [],
            }

        result = {
            "fy_summaries": {fy_label: _empty_fy_dict()},
            "trades": [],
            "open_holdings": [],
        }
        fy = result["fy_summaries"][fy_label]

        for row in rows:
            if len(row) <= max(idx.values()):
                continue
            scrip = row[idx["scrip"]]
            if not scrip:
                continue
            scrip = str(scrip).strip()
            qty = _num(row[idx["quantity"]])
            if qty <= 0:
                continue
            buy_val = _num(row[idx["buy_value"]])
            sell_val = _num(row[idx["sell_value"]])
            pnl = _num(row[idx["pnl"]]) if idx.get("pnl", -1) >= 0 else (sell_val - buy_val)
            charges = _num(row[idx["charges"]]) if idx.get("charges", -1) >= 0 else 0.0

            fy["equity_buy_value"] += buy_val
            fy["equity_sell_value"] += sell_val
            fy["equity_pnl"] += pnl

            buy_d = parse_date(row[idx["buy_date"]]) if idx.get("buy_date", -1) >= 0 else None
            sell_d = parse_date(row[idx["sell_date"]]) if idx.get("sell_date", -1) >= 0 else None
            isin_v = row[idx["isin"]] if idx.get("isin", -1) >= 0 else None

            if buy_d and sell_d and (sell_d - buy_d).days > 365:
                fy["equity_ltcg"] += pnl
            else:
                fy["equity_stcg"] += pnl

            result["trades"].append(Trade(
                scrip=scrip,
                isin=str(isin_v) if isin_v else None,
                quantity=qty,
                buy_date=buy_d,
                buy_value=buy_val,
                sell_date=sell_d,
                sell_value=sell_val,
                pnl=pnl,
                charges=charges,
                fy=fy_label,
                source_broker=self.name,
            ))

        return result


def _empty_fy_dict() -> dict:
    return {
        "equity_buy_value": 0.0, "equity_sell_value": 0.0, "equity_pnl": 0.0,
        "equity_stcg": 0.0, "equity_ltcg": 0.0,
        "equity_stamp_duty": 0.0, "equity_stt": 0.0, "equity_brokerage": 0.0,
        "equity_other_charges": 0.0, "equity_intraday_pnl": 0.0,
        "fno": {
            "options_turnover": 0.0, "options_pnl": 0.0,
            "futures_turnover": 0.0, "futures_pnl": 0.0,
            "stt": 0.0, "charges": 0.0, "brokerage": 0.0,
        },
        "dividend_income": 0.0,
        "open_holdings_cost": 0.0, "open_holdings_market_value": 0.0,
        "open_holdings_unrealised": 0.0,
        "open_holdings_st_unrealised": 0.0,
        "open_holdings_lt_unrealised": 0.0,
    }


# ---------- Column-mapping UI helpers ----------

# What the upload UI shows for the column-mapping step.
# `key` is canonical (used by the parser), `label` is shown to the user,
# `required` indicates must be mapped.
MAPPING_FIELDS: list[dict] = [
    {"key": "scrip",      "label": "Scrip / Symbol",        "required": True,  "example": "RELIANCE"},
    {"key": "isin",       "label": "ISIN",                   "required": False, "example": "INE002A01018"},
    {"key": "buy_date",   "label": "Buy date",               "required": False, "example": "2024-01-15"},
    {"key": "sell_date",  "label": "Sell date",              "required": False, "example": "2024-06-30"},
    {"key": "quantity",   "label": "Quantity",               "required": True,  "example": "100"},
    {"key": "buy_value",  "label": "Buy value (total)",      "required": True,  "example": "24500.00"},
    {"key": "sell_value", "label": "Sell value (total)",     "required": True,  "example": "28000.00"},
    {"key": "pnl",        "label": "Realised P&L (optional; computed if missing)",
                                                                    "required": False, "example": "3500.00"},
    {"key": "charges",    "label": "Charges (optional)",     "required": False, "example": "20.00"},
    {"key": "fy",         "label": "FY column (optional)",   "required": False, "example": "2024-25"},
]


def extract_headers(file: Path) -> list[str]:
    """Read just the header row of a file. Used by the upload UI to populate
    the column-mapping dropdowns before the user has mapped anything."""
    headers, _ = _read_rows(file)
    return headers
