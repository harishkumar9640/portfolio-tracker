"""Angel One SmartAPI 'Tax PNL' xlsx adapter.

The export has these sheets:
  - Equity+Bonds+SGB Trade Details   (closed trades + open holdings)
  - Derivatives Trade Details         (F&O summary)
  - Dividend Report                   (dividend income)

Sheet name detection is fuzzy (case-insensitive, partial-match) so a new
Angel One schema variant won't break the parser. Section labels inside
the equity sheet are also fuzzy-matched.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from pipeline.tax_pnl import (
    OpenHolding, Trade, _num, detect_fy_from_filename, parse_date,
)


def _norm(s) -> str:
    """Normalise a cell value for fuzzy comparison: lower, strip, collapse spaces."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _find_sheet(wb, *needles: str) -> str | None:
    """Find a sheet whose name (normalised) contains ALL needles."""
    for sheet_name in wb.sheetnames:
        norm = _norm(sheet_name)
        if all(_norm(n) in norm for n in needles):
            return sheet_name
    return None


def _find_row(rows, *needles: str, min_col: int = 0) -> tuple[int, list] | None:
    """Find the first row in `rows` where the first cell (normalised)
    contains ALL needles. Returns (index, row) or None.
    """
    needles_n = [_norm(n) for n in needles]
    for i, row in enumerate(rows):
        if not row:
            continue
        first = _norm(row[min_col]) if len(row) > min_col else ""
        if all(n in first for n in needles_n):
            return i, row
    return None


class AngelOneAdapter:
    name = "angel_one"

    def can_parse(self, file: Path) -> bool:
        """Angel One export has a sheet literally called
        'Equity+Bonds+SGB Trade Details' (or close to it)."""
        if file.suffix.lower() not in (".xlsx", ".xlsm"):
            return False
        try:
            wb = openpyxl.load_workbook(str(file), read_only=True, data_only=True)
            try:
                # Angel One uses + in the sheet name; some variants may use
                # spaces, &, or hyphens. We just look for "equity" + "bond"
                # + "trade".
                target = _find_sheet(wb, "equity", "bond", "trade")
                if target is None:
                    return False
                # Also check that the ISIN column appears in the equity sheet
                ws = wb[target]
                for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                    if row and _norm(row[0]) == "isin":
                        return True
                return False
            finally:
                wb.close()
        except Exception:
            return False

    def parse(self, file: Path) -> dict:
        wb = openpyxl.load_workbook(str(file), data_only=True)
        try:
            fy_label = detect_fy_from_filename(file.name)
            result = {
                "fy_summaries": {fy_label: _empty_fy_summary()},
                "trades": [],
                "open_holdings": [],
            }
            self._parse_equity_sheet(wb, fy_label, result)
            self._parse_derivatives_sheet(wb, fy_label, result)
            self._parse_dividend_sheet(wb, fy_label, result)
            return result
        finally:
            wb.close()

    # ---- private parsers ------------------------------------------------

    def _parse_equity_sheet(self, wb, fy_label: str, result: dict) -> None:
        sheet = _find_sheet(wb, "equity", "bond", "trade")
        if sheet is None:
            return
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        section = None
        fy_summary = result["fy_summaries"][fy_label]
        for row in rows:
            if not row or all(v is None or v == "" for v in row):
                continue
            first = _norm(row[0])
            if not first:
                continue
            # Summary rows
            if first == "net p&l" or first == "net p&l (a+b)":
                fy_summary["equity_pnl"] += _num(row[1])
            elif "ltcg" in first and "exclud" in first:
                fy_summary["equity_ltcg"] += _num(row[1])
            elif "stcg" in first and "exclud" in first:
                fy_summary["equity_stcg"] += _num(row[1])
            elif "intraday" in first and "speculative" in first:
                fy_summary["equity_intraday_pnl"] += _num(row[1])
            elif "total charges" in first:
                fy_summary["equity_other_charges"] += _num(row[1])
            elif first == "total stt":
                fy_summary["equity_stt"] += _num(row[1])
            elif "additional brokerage" in first:
                fy_summary["equity_brokerage"] += _num(row[1])
            # Section markers
            elif "intraday" in first:
                section = "intraday"; continue
            elif "delivery p&l" in first:
                section = "delivery"; continue
            elif "buyback" in first:
                section = "buyback"; continue
            elif "transfer" in first:
                section = "transfer"; continue
            elif "open sell" in first:
                section = "open_sell"; continue
            elif "open holding" in first:
                section = "open"; continue
            elif first == "isin":
                continue
            # Data row: must have an ISIN in column 0
            isin = row[0] if isinstance(row[0], str) and row[0].startswith("INE") else None
            if not isin:
                continue
            scrip = row[1] if len(row) > 1 and row[1] else None
            if not scrip:
                continue
            if section == "delivery" and len(row) > 12:
                qty = _num(row[2])
                buy_val = _num(row[6])
                sell_val = _num(row[8])
                charges = _num(row[10]) if len(row) > 10 else 0
                stt = _num(row[11]) if len(row) > 11 else 0
                pnl = _num(row[12]) if len(row) > 12 else (sell_val - buy_val)
                if qty > 0:
                    fy_summary["equity_buy_value"] += buy_val
                    fy_summary["equity_sell_value"] += sell_val
                    fy_summary["equity_stamp_duty"] += charges
                    fy_summary["equity_stt"] += stt
                    result["trades"].append(Trade(
                        scrip=str(scrip).strip(),
                        isin=isin,
                        quantity=qty,
                        buy_date=parse_date(row[3]),
                        buy_value=buy_val,
                        sell_date=parse_date(row[4]),
                        sell_value=sell_val,
                        pnl=pnl,
                        charges=charges,
                        stt=stt,
                        fy=fy_label,
                        source_broker=self.name,
                    ))
            elif section == "open" and len(row) > 10:
                qty = _num(row[2])
                buy_val = _num(row[4])
                closing = _num(row[7]) if len(row) > 7 else 0
                st_un = _num(row[9]) if len(row) > 9 else 0
                lt_un = _num(row[10]) if len(row) > 10 else 0
                if qty > 0 and buy_val > 1000:
                    fy_summary["open_holdings_cost"] += buy_val
                    fy_summary["open_holdings_market_value"] += closing * qty
                    fy_summary["open_holdings_st_unrealised"] += st_un
                    fy_summary["open_holdings_lt_unrealised"] += lt_un
                    result["open_holdings"].append(OpenHolding(
                        scrip=str(scrip).strip(),
                        isin=isin,
                        quantity=qty,
                        buy_value=buy_val,
                        current_value=closing * qty,
                        unrealised=(closing * qty) - buy_val,
                        st_unrealised=st_un,
                        lt_unrealised=lt_un,
                        fy=fy_label,
                    ))

    def _parse_derivatives_sheet(self, wb, fy_label: str, result: dict) -> None:
        sheet = _find_sheet(wb, "derivative")
        if sheet is None:
            return
        ws = wb[sheet]
        fno = result["fy_summaries"][fy_label].setdefault("fno", _empty_fno())
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if not row or not row[0]:
                continue
            first = _norm(row[0])
            if first == "futures turnover":
                fno["futures_turnover"] += _num(row[1])
            elif first == "options turnover":
                fno["options_turnover"] += _num(row[1])
            elif "futures p&l" in first:
                fno["futures_pnl"] += _num(row[1])
            elif "options p&l" in first:
                fno["options_pnl"] += _num(row[1])
            elif "total charges" in first:
                fno["charges"] += _num(row[1])
            elif first == "total stt":
                fno["stt"] += _num(row[1])

    def _parse_dividend_sheet(self, wb, fy_label: str, result: dict) -> None:
        sheet = _find_sheet(wb, "dividend")
        if sheet is None:
            return
        ws = wb[sheet]
        fy_summary = result["fy_summaries"][fy_label]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row and len(row) > 5 and isinstance(row[5], (int, float)):
                fy_summary["dividend_income"] += float(row[5])


def _empty_fy_summary() -> dict:
    return {
        "equity_buy_value": 0.0,
        "equity_sell_value": 0.0,
        "equity_pnl": 0.0,
        "equity_stcg": 0.0,
        "equity_ltcg": 0.0,
        "equity_stamp_duty": 0.0,
        "equity_stt": 0.0,
        "equity_brokerage": 0.0,
        "equity_other_charges": 0.0,
        "equity_intraday_pnl": 0.0,
        "fno": _empty_fno(),
        "dividend_income": 0.0,
        "open_holdings_cost": 0.0,
        "open_holdings_market_value": 0.0,
        "open_holdings_unrealised": 0.0,
        "open_holdings_st_unrealised": 0.0,
        "open_holdings_lt_unrealised": 0.0,
    }


def _empty_fno() -> dict:
    return {
        "options_turnover": 0.0,
        "options_pnl": 0.0,
        "futures_turnover": 0.0,
        "futures_pnl": 0.0,
        "stt": 0.0,
        "charges": 0.0,
        "brokerage": 0.0,
    }
