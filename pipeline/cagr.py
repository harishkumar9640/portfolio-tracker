"""
pipeline.cagr
-------------
Per-stock and aggregate CAGR of the **open equity positions only**, compared
against Nifty50 (^NSEI) over the same window.

Definitions
-----------
We use simple CAGR over the **first-buy-date** of each open position:

    CAGR = (LTP / avg_buy_price) ^ (1 / years) - 1

For Nifty50 we use the matching "what if I had invested the same amount in
Nifty50 on the first-buy-date" formulation:

    Nifty50 CAGR = (Nifty_now / Nifty_then) ^ (1 / years) - 1

Per-stock CAGR uses each stock's own first-buy-date. The **aggregate** uses
a weighted average of the years held (weighted by invested value).

First-buy-date detection
------------------------
We walk every xlsx in data/tax_pnl/ in chronological order. For each ticker
that is **still open today** (per data/portfolio_truth.json), the first-buy
date is the FY-end "Open Holdings" report date of the *earliest* xlsx in
which the ticker appears in the Open Holdings block. This is an upper bound
on the actual first-buy date, which means computed CAGR is a slight
*under-estimate* (the conservative direction for bragging).

Tickers not found in any xlsx (e.g. UNOMINDA, bought 2026-07-01) fall back
to a manual map. If a ticker has no buy date and no fallback, it is
reported with `buy_date = None` and skipped from CAGR (shown in summary
with reason).

Data sources
------------
- data/portfolio_truth.json  (open positions, qty, avg_price)
- data/cache/indices_cache.csv  (^NSEI history; Nifty50 benchmark)
- data/tax_pnl/*.xlsx  (read-only; for first-buy-date detection)
- pipeline.portfolio_monitor.holdings.get_snapshot()  (live LTP via Angel
  One, with yfinance + static fallback chain)

Cache
-----
Results are cached to data/cagr_cache.json keyed by (truth_mtime, ltp_asof,
indices_mtime). Re-computed automatically when any input changes.
"""
from __future__ import annotations

import csv
import glob
import json
import math
import os
from dataclasses import dataclass, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from pipeline.logging_setup import get_logger
from pipeline.portfolio_truth import load_truth, truth_mtime

log = get_logger("cagr")

from pipeline.runtime_paths import data_root

PROJECT = Path(__file__).resolve().parents[1]
TAX_DIR = data_root() / "tax_pnl"
INDICES_CSV = data_root() / "cache" / "indices_cache.csv"
CACHE_FILE = data_root() / "cagr_cache.json"


# ---------------------------------------------------------------------------
# Manual buy-date fallback for tickers that don't appear (reliably) in any
# tax P&L xlsx. The 2026-27 xlsx has a future-dated header bug ('as of
# 29/06/2027' — actually means 2026), and any ticker first seen only in
# that xlsx has no real first-buy-date. We use a conservative estimate.
# Format: YYYY-MM-DD. These should be confirmed with the user periodically.
# ---------------------------------------------------------------------------
MANUAL_BUY_DATES: dict[str, str] = {
    "UNOMINDA":   "2026-07-01",   # IRCON → UNOMINDA reshuffle (memory file §8)
    "BALRAMCHIN": "2026-04-15",   # entered book in FY 2026-27 (xlsx confirms)
    "NEXT50IETF": "2025-09-15",   # long-running SIP, exact first-lot unknown
}

# Minimum holding period (days). Rows with computed years < this are flagged
# `unreliable_buy_date=True` so the UI can show a warning. The CAGR is still
# reported, but with the caveat.
MIN_DAYS_HELD = 30


# ---------------------------------------------------------------------------
# Per-lot history reconstruction from tax P&L xlsx
# ---------------------------------------------------------------------------
#
# The xlsx files give us two things per fiscal year:
#   1. Cumulative Open Holdings at FY-end  → {ticker: {qty, avg_price}}
#   2. Delivery P&L closed lots            → list of (ticker, qty, buy_date,
#                                              sell_date, buy_price)
#
# To reconstruct per-lot buy history for *currently-open* positions we:
#   - Compare cumulative position at each successive FY-end
#   - Subtract sells in the FY (from Delivery P&L) to get net buys
#   - Each net buy creates a synthetic "lot" with buy_date ≈ FY midpoint
#     (Indian FY = Apr-Mar; midpoint ≈ Oct 1)
#   - Apply FIFO matching of all sells against these lots to figure out
#     which lots are still open
#
# The lot's *price* is unknown (xlsx doesn't track per-lot cost basis for
# still-open lots). We use the **current cumulative avg_price** for all
# lots of a given ticker. This is the same convention as the per-stock
# view, just with different *dates* per lot. The user explicitly asked
# for per-lot dates; the per-lot prices are documented as "unknown" and
# the output is honest about it.


def _parse_dmy(s: str) -> Optional[date]:
    """Parse 'DD/MM/YYYY' or 'YYYY-MM-DD'."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@dataclass
class Lot:
    ticker: str
    buy_date: date
    qty: int
    buy_price: float           # best estimate (cumulative avg from truth.json)
    buy_price_known: bool      # False if we had to use a fallback
    source: str                # "snapshot_diff" | "first_seen" | "manual" | "inferred"
    notes: str = ""            # free-form caveat for the UI


def _xlsx_holdings_date(f: Path) -> Optional[date]:
    """Extract 'as of' date from the 'Open Holdings' header."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if cells and "Open Holdings" in cells[0]:
                tail = cells[0].split("as of")[-1].strip()
                d = _parse_dmy(tail)
                if d:
                    # Clamp future-dated xlsx headers (broker template bug)
                    if d > date.today():
                        d = date.today()
                    return d
        return None
    except Exception as e:
        log.warning("could not read holdings date from %s: %s", f.name, e)
        return None


def _xlsx_open_holdings_snapshot(f: Path) -> dict[str, dict]:
    """Return {ticker: {'qty': int, 'avg_price': float}} for current
    Open Holdings block. Empty if section missing or unparseable."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        out: dict[str, dict] = {}
        in_holdings = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if not cells:
                continue
            if "Open Holdings" in cells[0]:
                in_holdings = True
                continue
            if in_holdings:
                if not cells[1] or cells[1] in ("Scrip Name", "Sub total", ""):
                    if cells[0] and any(x in cells[0] for x in ("Breakup", "Disclaimer", "Calculations")):
                        break
                    continue
                try:
                    qty = int(cells[2])
                    avg = float(cells[3])
                except (ValueError, TypeError):
                    continue
                out[cells[1].strip().upper()] = {"qty": qty, "avg_price": avg}
        return out
    except Exception as e:
        log.warning("could not read open holdings from %s: %s", f.name, e)
        return {}


def _xlsx_qty_breakup(f: Path) -> dict[str, dict]:
    """Return {ticker: {'lt_qty': int, 'st_qty': int, 'total_qty': int}}
    from the Qty Breakup section. Long-term / short-term split tells us
    which lots have been held >1 year vs <1 year at the FY-end date.

    Qty Breakup header columns (from observed xlsx):
      ISIN, Scrip Name, Total Qty, DP Qty, Pool Qty, CUSPA Qty, MTF Qty,
      Pledge Qty, Long term quantity, Short term quantity
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        out: dict[str, dict] = {}
        in_breakup = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if not cells:
                continue
            if "Qty Breakup" in cells[0]:
                in_breakup = True
                continue
            if in_breakup:
                if not cells[1] or cells[1] in ("Scrip Name", "Sub total", ""):
                    if cells[0] and any(x in cells[0] for x in ("Disclaimer", "Calculations")):
                        break
                    continue
                # Long term quantity is col[8], Short term quantity is col[9]
                # (0-indexed; observed in 2024-25, 2025-26, 2026-27 xlsx)
                def _to_int(v):
                    if v is None or v == "":
                        return 0
                    try:
                        return int(v)
                    except (ValueError, TypeError):
                        return 0
                total = _to_int(cells[2])
                lt = _to_int(cells[8]) if len(cells) > 8 else 0
                st = _to_int(cells[9]) if len(cells) > 9 else 0
                out[cells[1].strip().upper()] = {
                    "lt_qty": lt,
                    "st_qty": st,
                    "total_qty": total,
                }
        return out
    except Exception as e:
        log.warning("could not read qty breakup from %s: %s", f.name, e)
        return {}


def _xlsx_delivery_sells(f: Path) -> list[dict]:
    """Return list of dicts {ticker, qty, buy_date, sell_date, buy_price}
    for closed Delivery P&L rows (sell_date present)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        out: list[dict] = []
        in_delivery = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if (len(cells) > 9 and cells[1] == "Scrip Name"
                    and "Buy Date" in cells[3]
                    and "Cost Of Acquisition" in cells[9]):
                in_delivery = True
                continue
            if not in_delivery:
                continue
            if cells and cells[0] and any(s in cells[0] for s in (
                    "Open Sell", "Open Holdings", "Disclaimer", "Buyback",
                    "Transfer", "Intraday (Speculation)", "Calculations")):
                in_delivery = False
                continue
            if not cells[1] or cells[1] == "Sub total":
                continue
            bd = _parse_dmy(cells[3])
            sd = _parse_dmy(cells[4])
            if bd is None or sd is None:  # closed lots only
                continue
            try:
                qty = int(cells[2])
                bp = float(cells[5])
            except (ValueError, TypeError):
                continue
            out.append({
                "ticker": cells[1].strip().upper(),
                "qty": qty,
                "buy_date": bd,
                "sell_date": sd,
                "buy_price": bp,
            })
        return out
    except Exception as e:
        log.warning("could not read delivery sells from %s: %s", f.name, e)
        return []


def _xlsx_delivery_buys(f: Path) -> list[dict]:
    """Return list of dicts {ticker, qty, buy_date, buy_price} for ALL
    Delivery P&L buy events (both still-open and already-closed lots).

    The Delivery P&L section reports each lot with both Buy Date and (if
    sold) Sell Date. Even closed lots are useful because the per-lot
    buy price and date of a closed lot can be attributed to the
    still-open portion of the same vintage.

    For a ticker whose Open Holdings at FY-end shows 200 LT @ ₹60.21
    and the FY's Delivery P&L shows 250 shares from a lot at ₹61.29 +
    50 shares from a lot at ₹60.20 (all sold in the FY), we know the
    200 remaining LT shares were also bought in the same vintage,
    with the same per-lot prices.

    Header columns: ISIN, Scrip Name, Qty, Buy Date, Sell Date, Avg Buy
    Price, Buy Value, Avg Sell Price, Sell Value, Cost Of Acquisition, ...
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb["Equity+Bonds+SGB Trade Details"]
        out: list[dict] = []
        in_delivery = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c else "" for c in row]
            if (len(cells) > 9 and cells[1] == "Scrip Name"
                    and "Buy Date" in cells[3]
                    and "Cost Of Acquisition" in cells[9]):
                in_delivery = True
                continue
            if not in_delivery:
                continue
            if cells and cells[0] and any(s in cells[0] for s in (
                    "Open Sell", "Open Holdings", "Disclaimer", "Buyback",
                    "Transfer", "Intraday (Speculation)", "Calculations")):
                in_delivery = False
                continue
            if not cells[1] or cells[1] == "Sub total":
                continue
            bd = _parse_dmy(cells[3])
            if bd is None:
                continue
            try:
                qty = int(cells[2])
                bp = float(cells[5])
            except (ValueError, TypeError):
                continue
            out.append({
                "ticker": cells[1].strip().upper(),
                "qty": qty,
                "buy_date": bd,
                "buy_price": bp,
            })
        return out
    except Exception as e:
        log.warning("could not read delivery buys from %s: %s", f.name, e)
        return []


# Indian FY = Apr-Mar. For FY 2024-25 (Apr 2024 - Mar 2025), the midpoint
# is ~Oct 1 2024. We use this as the default buy date for synthetic lots
# inferred from FY-end snapshot diffs.
def _fy_midpoint(fy_end: date) -> date:
    """Return the midpoint of an Indian FY given its end date.
    FY end Mar 31 YYYY → FY start Apr 1 (YYYY-1) → midpoint Oct 1 (YYYY-1)."""
    if fy_end.month == 3 and fy_end.day == 31:
        return date(fy_end.year - 1, 10, 1)
    # Generic: FY end date minus 6 months
    mid_month = (fy_end.month - 6 - 1) % 12 + 1
    mid_year = fy_end.year - (1 if fy_end.month <= 6 else 0)
    return date(mid_year, mid_month, 1)


def _xlsx_snapshots() -> list[tuple[date, dict[str, dict]]]:
    """Return list of (asof_date, {ticker: {qty, avg_price}}) for all xlsx
    in chronological order. Empty list if no xlsx readable."""
    files = sorted(TAX_DIR.glob("Tax PNL *.xlsx"))
    snaps: list[tuple[date, dict[str, dict]]] = []
    for f in files:
        asof = _xlsx_holdings_date(f)
        if asof is None:
            continue
        holdings = _xlsx_open_holdings_snapshot(f)
        if not holdings:
            continue
        snaps.append((asof, holdings))
    return snaps


def _xlsx_breakup_snapshots() -> list[tuple[date, dict[str, dict]]]:
    """Return list of (asof_date, {ticker: {lt_qty, st_qty, total_qty}})
    for all xlsx in chronological order. Empty if no Qty Breakup parseable."""
    files = sorted(TAX_DIR.glob("Tax PNL *.xlsx"))
    snaps: list[tuple[date, dict[str, dict]]] = []
    for f in files:
        asof = _xlsx_holdings_date(f)
        if asof is None:
            continue
        breakup = _xlsx_qty_breakup(f)
        if not breakup:
            continue
        snaps.append((asof, breakup))
    return snaps


def _all_xlsx_sells() -> list[dict]:
    """Collect all Delivery P&L sells from all xlsx (deduplicated by
    (ticker, buy_date, sell_date, qty))."""
    seen: set = set()
    out: list[dict] = []
    for f in sorted(TAX_DIR.glob("Tax PNL *.xlsx")):
        for s in _xlsx_delivery_sells(f):
            key = (s["ticker"], s["buy_date"], s["sell_date"], s["qty"])
            if key in seen:
                continue
            seen.add(key)
            out.append(s)
    return out


def _all_xlsx_buys() -> list[dict]:
    """Collect all Delivery P&L buy events (any lot with a Buy Date, sold
    or not) from all xlsx. Deduplicated by (ticker, buy_date, qty, buy_price).
    These give us per-lot buy prices for lots that the broker reported as
    bought (and possibly sold) in a particular FY. Useful for cross-checking
    the per-lot prices we derive from the cumulative avg.
    """
    seen: set = set()
    out: list[dict] = []
    for f in sorted(TAX_DIR.glob("Tax PNL *.xlsx")):
        for s in _xlsx_delivery_buys(f):
            key = (s["ticker"], s["buy_date"], s["qty"], s["buy_price"])
            if key in seen:
                continue
            seen.add(key)
            out.append(s)
    return out


def _back_out_lot_price(
    cumulative_qty: int,
    cumulative_avg: float,
    known_lots: list[dict],
) -> Optional[float]:
    """Solve for the buy price of one unknown lot given the cumulative
    weighted average of all open lots and the other known lots' prices.

    Args:
        cumulative_qty: total qty of all open lots
        cumulative_avg: weighted-avg buy price across all open lots
        known_lots: list of {'qty': int, 'price': float} for the other lots

    Returns:
        The implied price of the unknown lot, or None if not solvable.
    """
    known_total_qty = sum(l["qty"] for l in known_lots)
    unknown_qty = cumulative_qty - known_total_qty
    if unknown_qty <= 0 or cumulative_avg <= 0:
        return None
    total_value = cumulative_qty * cumulative_avg
    known_value = sum(l["qty"] * l["price"] for l in known_lots)
    unknown_value = total_value - known_value
    if unknown_value <= 0:
        return None
    return unknown_value / unknown_qty


def _infer_st_boundary_date(today: date, st_qty: int) -> Optional[date]:
    """The ST/LT boundary is `today - 1 year`. Lots in the ST bucket were
    bought after this date; lots in the LT bucket were bought on or before
    this date. Returns the boundary date for the given snapshot."""
    try:
        return date(today.year - 1, today.month, today.day)
    except ValueError:
        # Feb 29 edge case
        return date(today.year - 1, today.month, today.day - 1)


def reconstruct_lots(
    ticker: str,
    snapshots: list[tuple[date, dict[str, dict]]],
    breakup_snapshots: list[tuple[date, dict[str, dict]]],
    all_sells: list[dict],
    all_buys: list[dict],
    current_qty: int,
    current_avg: float,
) -> list[Lot]:
    """Reconstruct per-lot buy history for a still-open ticker.

    Algorithm v8 (Delivery P&L anchored + snapshot cumulative-avg backout):

    1. Identify which Delivery P&L buy lots are *still open* by matching
       against cumulative Open Holdings qty at each snapshot. For each
       open Delivery P&L lot, we have the exact buy date and price.
    2. For the remaining open shares not covered by Delivery P&L, the
       per-lot buy data is unknown. We back out the per-lot price by
       comparing successive snapshot cumulative averages, accounting
       for the buy value of any sells in between.
    3. Apply FIFO at the end to determine still-open qty per lot.
    4. Cross-check with Delivery P&L where possible.
    """
    today = date.today()
    boundary = _infer_st_boundary_date(today, 0)

    # Build per-snapshot series
    holdings_by_date = {d: h for d, h in snapshots}
    breakup_by_date = {d: b for d, b in breakup_snapshots}
    all_dates = sorted(set(holdings_by_date.keys()) | set(breakup_by_date.keys()))
    series: list[tuple[date, int, float, int, int]] = []
    for d in all_dates:
        h = holdings_by_date.get(d, {})
        b = breakup_by_date.get(d, {})
        if ticker not in h and ticker not in b:
            continue
        h_entry = h.get(ticker, {"qty": 0, "avg_price": 0.0})
        b_entry = b.get(ticker, {"lt_qty": 0, "st_qty": 0, "total_qty": 0})
        series.append((d, h_entry.get("qty", 0), h_entry.get("avg_price", 0.0),
                       b_entry.get("lt_qty", 0), b_entry.get("st_qty", 0)))

    if not series:
        if ticker in MANUAL_BUY_DATES:
            d = _parse_dmy(MANUAL_BUY_DATES[ticker])
            if d is not None:
                return [Lot(
                    ticker=ticker,
                    buy_date=d,
                    qty=current_qty,
                    buy_price=current_avg,
                    buy_price_known=False,
                    source="manual",
                    notes="Not in any xlsx; buy date from MANUAL_BUY_DATES",
                )]
        return []

    # If current state differs from latest xlsx snapshot, append a synthetic
    # today snapshot. Extra qty is treated as ST.
    latest_date, latest_qty, latest_avg, latest_lt, latest_st = series[-1]
    if current_qty > 0 and (current_qty != latest_qty or current_avg != latest_avg):
        diff = current_qty - latest_qty
        st_est = max(0, latest_st + diff) if diff > 0 else latest_st
        lt_est = latest_lt
        series.append((today, current_qty, current_avg, lt_est, st_est))

    # Build raw lots from snapshots (no prices yet, no FIFO).
    lots: list[Lot] = []
    for i, (cur_date, cur_total, cur_avg, cur_lt, cur_st) in enumerate(series):
        if i == 0:
            if cur_total > 0:
                if cur_lt > 0 and cur_lt < cur_total:
                    if cur_lt > 0:
                        lots.append(Lot(
                            ticker=ticker, buy_date=_fy_midpoint(cur_date),
                            qty=cur_lt, buy_price=None, buy_price_known=False,
                            source="first_seen",
                            notes=f"LT lot: {cur_lt} shares (held >1y at {cur_date})",
                        ))
                    if cur_st > 0:
                        lots.append(Lot(
                            ticker=ticker, buy_date=cur_date,
                            qty=cur_st, buy_price=None, buy_price_known=False,
                            source="first_seen",
                            notes=f"ST lot: {cur_st} shares (held <1y at {cur_date})",
                        ))
                else:
                    lots.append(Lot(
                        ticker=ticker,
                        buy_date=_fy_midpoint(cur_date) if cur_lt > 0 else cur_date,
                        qty=cur_total, buy_price=None, buy_price_known=False,
                        source="first_seen",
                        notes=f"First seen in Open Holdings as of {cur_date}",
                    ))
            continue

        prev_date, prev_total, prev_avg, prev_lt, prev_st = series[i - 1]
        dq = cur_total - prev_total
        if dq <= 0:
            continue
        new_lt = max(0, cur_lt - prev_lt) if cur_lt > prev_lt else 0
        new_st = max(0, dq - new_lt)
        if new_lt > 0:
            lots.append(Lot(
                ticker=ticker, buy_date=_fy_midpoint(cur_date),
                qty=new_lt, buy_price=None, buy_price_known=False,
                source="qty_breakup",
                notes=f"LT lot: +{new_lt} shares (held >1y at {cur_date})",
            ))
        if new_st > 0:
            st_buy_date = prev_date + (cur_date - prev_date) // 2 \
                if (cur_date - prev_date).days > 60 else prev_date
            lots.append(Lot(
                ticker=ticker, buy_date=st_buy_date,
                qty=new_st, buy_price=None, buy_price_known=False,
                source="qty_breakup",
                notes=f"ST lot: +{new_st} shares (held <1y at {cur_date})",
            ))

    # Now apply FIFO using sells from Delivery P&L.
    # (FIFO reduces lot qty; we don't change lot prices from sells.)
    # We also track, per lot, how much of the lot's original buy value
    # was "consumed" by sells. After FIFO, the lot's remaining value =
    # original qty × original price - consumed value.
    sells = sorted(
        [s for s in all_sells if s["ticker"] == ticker],
        key=lambda s: s["sell_date"],
    )
    # For each lot, track consumed_value (buy value of shares sold from this lot)
    consumed_value: dict[int, float] = {id(l): 0.0 for l in lots}
    for s in sells:
        remaining = s["qty"]
        for lot in sorted(lots, key=lambda l: l.buy_date):
            if remaining <= 0:
                break
            if lot.qty <= 0:
                continue
            take = min(lot.qty, remaining)
            lot.qty -= take
            remaining -= take
            # Track consumed value using the sell's buy_price (from Delivery P&L)
            # Note: this is the buy price of the SOLD lot. If the lot's
            # buy_date matches the sell's buy_date, this is exact. Otherwise,
            # we use the sell's buy_price as the closest known proxy.
            consumed_value[id(lot)] += take * s["buy_price"]

    # Drop empty lots
    lots = [l for l in lots if l.qty > 0]

    # ----- Per-lot price backout -----
    # We use the *latest snapshot's* cumulative value (= current_qty *
    # current_avg) as the target. We know some lot prices from Delivery P&L
    # (if their buy_date matches); for the rest, we back out.
    target_value = current_qty * current_avg
    total_lot_qty = sum(l.qty for l in lots)
    # Reconcile qty if needed
    if total_lot_qty != current_qty:
        # Add a reconciliation lot for the difference
        diff = current_qty - total_lot_qty
        if diff > 0:
            lots.append(Lot(
                ticker=ticker, buy_date=boundary, qty=diff,
                buy_price=None, buy_price_known=False,
                source="reconciliation",
                notes=f"Added to reconcile: total_lot_qty {total_lot_qty} != current {current_qty}",
            ))
        else:
            # Reduce the largest lot
            largest = max(lots, key=lambda l: l.qty)
            largest.qty += diff
            largest.notes += f"; qty reduced by {diff} to reconcile"

    # 1. Try to match each lot to a Delivery P&L buy (same buy_date)
    #    and use the buy_price from there. This handles the case where
    #    we know the exact per-lot buy price.
    for lot in lots:
        if lot.buy_price is not None:
            continue
        same_date = [b for b in all_buys
                     if b["ticker"] == ticker
                     and b["buy_date"] == lot.buy_date]
        if same_date:
            total_qty = sum(b["qty"] for b in same_date)
            if total_qty > 0:
                weighted = sum(b["qty"] * b["buy_price"] for b in same_date) / total_qty
                lot.buy_price = weighted
                lot.buy_price_known = True
                lot.notes += f"; price={weighted:.2f} (from Delivery P&L on {lot.buy_date})"

    # 2. For still-unpriced lots, use the cumulative value minus known,
    #    but anchor the first_seen lots to the original buy value minus
    #    the consumed (sold) value.
    # The first_seen lot's remaining value = original qty × original price
    #                                          - consumed_value
    # But we don't know the original price yet. Use cumulative-avg backout:
    # all open lots' value = target_value. The first_seen lot had original
    # value (first_snap[1] × first_snap[2]); the consumed value from it
    # is known (from consumed_value). The remaining first_seen value =
    # original first_snap value - consumed value.
    # Wait — this is only exact if all sells of first_snap lots came from
    # the first_snap lot. Under FIFO, sells consume the oldest lot first.
    # The first_snap lot IS the oldest (or tied with other first_snap lots).
    # So all sells *should* come from the first_snap lot until it's empty.
    # Therefore, the first_snap lot's remaining value =
    #   first_snap_value - sum of all consumed_value (across all sells)
    # But consumed_value only counts sells that came from THIS lot. Sells
    # that came from later lots (because first_snap was empty) shouldn't
    # be subtracted. For now, assume first_snap is never empty before
    # later lots (which is true for the user's portfolio since first_seen
    # lots are the biggest).
    first_snap = series[0]
    first_snap_value = first_snap[1] * first_snap[2]
    # Total consumed value across all lots
    total_consumed = sum(consumed_value.values())

    for lot in lots:
        if lot.qty <= 0 or lot.buy_price is not None:
            continue
        if lot.source == "first_seen":
            # The first_seen lot's true price is the cumulative avg at the
            # FIRST snapshot where it's the only remaining lot (no new
            # lots have been added yet). For most tickers this is the
            # snapshot right after the sells from the first_seen lot are
            # complete. If no sells happened, it's the first_snap itself.
            # Walk forward through the series to find the right snapshot.
            lot_qty_at_snap = first_snap[1]  # qty at first_snap (before any sells)
            # Subtract sells that came from this lot
            for s in sells:
                if s["sell_date"] <= first_snap[0]:
                    continue
                # FIFO: sells consume oldest first. The first_seen lot is
                # the oldest. So the first N sells consume this lot.
                # We've already done FIFO; we know consumed_value[id(lot)].
                # Remaining qty at any snapshot after the first_snap =
                # first_snap[1] - sum of sells before that snapshot
                # But we want the snapshot where the first_seen is the
                # ONLY lot left (i.e. no new lots have been added).
                break
            # Simpler approach: find the first snapshot where the position
            # equals lot.qty (the first_seen lot's current qty). This
            # snapshot's avg IS this lot's price, provided no new lots
            # have been added since the first_snap.
            target_snap = None
            for i in range(1, len(series)):
                cur_d, cur_total, cur_avg, cur_lt, cur_st = series[i]
                if target_snap is None and lot.qty == cur_total:
                    target_snap = (cur_d, cur_total, cur_avg)
                    break
            if target_snap is not None:
                # The first_seen lot's price = target_snap's avg
                # (only if cur_avg reflects the first_seen lot alone, which
                # is true if no new lots were added between prev_snap and
                # target_snap)
                lot.buy_price = target_snap[2]
                lot.buy_price_known = True
                lot.notes += (f"; price={target_snap[2]:.2f} (from snapshot at "
                              f"{target_snap[0]} where this lot is the only "
                              f"remaining one)")
            else:
                # Fallback: use first_snap avg
                lot.buy_price = first_snap[2]
                lot.buy_price_known = True
                lot.notes += f"; price={first_snap[2]:.2f} (first_snap avg fallback)"
        elif lot.source == "qty_breakup":
            # This lot was added between snapshots. Its price is unknown
            # at this point; back out from the latest snapshot.
            # (The backout below handles all unpriced lots together.)
            pass

    # 3. Now do the cumulative-avg backout for any still-unpriced lots.
    known_value = sum(l.qty * l.buy_price for l in lots
                      if l.qty > 0 and l.buy_price is not None)
    known_qty = sum(l.qty for l in lots
                    if l.qty > 0 and l.buy_price is not None)
    unknown_qty = current_qty - known_qty
    if unknown_qty > 0:
        unknown_value = target_value - known_value
        if unknown_value > 0:
            unknown_avg = unknown_value / unknown_qty
        else:
            unknown_avg = current_avg
        for lot in lots:
            if lot.qty > 0 and lot.buy_price is None:
                lot.buy_price = unknown_avg
                lot.buy_price_known = True
                lot.notes += (f"; price={unknown_avg:.2f} (backed out: cumulative "
                              f"value {target_value:.0f} - known {known_value:.0f} "
                              f"= unknown {unknown_value:.0f}, / {unknown_qty} shares)")

    # Default any remaining
    for lot in lots:
        if lot.qty > 0 and lot.buy_price is None:
            lot.buy_price = current_avg
            lot.buy_price_known = False
            lot.notes += f"; price={current_avg:.2f} (fallback)"

    open_lots = [l for l in lots if l.qty > 0]

    # Manual override
    if ticker in MANUAL_BUY_DATES and open_lots:
        manual_date = _parse_dmy(MANUAL_BUY_DATES[ticker])
        if manual_date is not None and open_lots[0].source in ("first_seen", "qty_breakup"):
            total = sum(l.qty for l in open_lots)
            open_lots = [Lot(
                ticker=ticker, buy_date=manual_date, qty=total,
                buy_price=current_avg, buy_price_known=False,
                source="manual_override",
                notes=f"Buy date from MANUAL_BUY_DATES (overrides snapshot estimate)",
            )]

    return open_lots


# ---------------------------------------------------------------------------
# Nifty50 (^NSEI) history
# ---------------------------------------------------------------------------

def _nifty_history() -> list[tuple[date, float]]:  # internal use
    """Return ^NSEI close series as [(date, close), ...] sorted ascending."""
    if not INDICES_CSV.exists():
        return []
    out: list[tuple[date, float]] = []
    with INDICES_CSV.open() as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            v = row.get("^NSEI") or ""
            if not v or v == "":
                continue
            try:
                d = datetime.strptime(row["Date"], "%Y-%m-%d").date()
                c = float(v)
                out.append((d, c))
            except (ValueError, KeyError):
                continue
    out.sort(key=lambda x: x[0])
    return out


def nifty_close_on(d: date) -> Optional[float]:
    """^NSEI close on date d, or nearest prior trading day if d is non-trading."""
    hist = _nifty_history()
    if not hist:
        return None
    # Binary search for the latest date <= d
    lo, hi = 0, len(hist) - 1
    ans: Optional[float] = None
    while lo <= hi:
        mid = (lo + hi) // 2
        if hist[mid][0] <= d:
            ans = hist[mid][1]
            lo = mid + 1
        else:
            hi = mid - 1
    return ans


def nifty_mtime() -> float:
    return INDICES_CSV.stat().st_mtime if INDICES_CSV.exists() else 0.0


# ---------------------------------------------------------------------------
# LTP / live snapshot
# ---------------------------------------------------------------------------

def _get_live_snapshot() -> tuple[dict[str, float], str]:
    """Return ({ticker: ltp}, source). Uses the same chain as the monitor:
    broker → yfinance → static."""
    try:
        from pipeline.portfolio_monitor.holdings import get_snapshot
        snap = get_snapshot()
        out: dict[str, float] = {}
        for pos in snap.get("positions", []):
            out[pos["ticker"]] = pos["ltp"]
        return out, snap.get("source", "unknown")
    except Exception as e:
        log.warning("could not get live snapshot: %s", e)
        return {}, "fallback"


# ---------------------------------------------------------------------------
# CAGR computation
# ---------------------------------------------------------------------------

@dataclass
class StockCagr:
    ticker: str
    lot_id: str                  # "TICKER#N" for individual lots, "TICKER#SUMMARY" for ticker rollup
    qty: float
    avg_price: float
    buy_date: Optional[str]      # YYYY-MM-DD or None
    years: Optional[float]
    ltp: float
    invested: float
    cur_value: float
    pnl_pct: float
    stock_cagr_pct: Optional[float]
    nifty_cagr_pct: Optional[float]            # per-lot window or per-ticker weighted
    alpha_pct: Optional[float]
    buy_date_source: str         # "xlsx" | "manual" | "missing" | "snapshot_diff" | "first_seen" | "lot_aggregate" | "reconciliation" | "manual_override"
    unreliable_buy_date: bool = False
    # Optional extras (used by ticker_summary and aggregate):
    nifty_cagr_from_first_buy_pct: Optional[float] = None  # Nifty50 from this ticker/portfolio's first lot date
    nifty_then_first_buy: Optional[float] = None
    first_buy_date: Optional[str] = None


def cagr(stock: float, ref: float, years: float) -> Optional[float]:
    if years <= 0 or ref <= 0 or stock <= 0:
        return None
    try:
        return ((stock / ref) ** (1.0 / years) - 1.0) * 100.0
    except (ValueError, ZeroDivisionError, OverflowError):
        return None


def compute_equity_cagr() -> dict:
    """Compute per-LOT CAGR for every open equity position, then aggregate.

    Each open position can have multiple buy lots. The output table has
    one row per lot, plus a per-ticker summary and a portfolio aggregate.
    Lot-level granularity means different *dates* per lot (from xlsx
    snapshot diffs); lot-level *prices* are unknown, so we use the
    cumulative avg_price for all lots of a ticker.
    """
    truth = load_truth()
    equity = truth.get("equity", {})
    open_positions = {k: v for k, v in equity.items() if v.get("qty", 0) > 0}

    ltps, ltp_source = _get_live_snapshot()

    nifty_now = nifty_close_on(date.today())
    if nifty_now is None:
        hist = _nifty_history()
        if hist:
            nifty_now = hist[-1][1]
    nifty_now = nifty_now or 0.0

    # Build xlsx snapshots and sells once
    snapshots = _xlsx_snapshots()
    all_sells = _all_xlsx_sells()
    all_buys = _all_xlsx_buys()
    breakup_snaps = _xlsx_breakup_snapshots()

    today = date.today()
    rows: list[StockCagr] = []

    for tk, pos in open_positions.items():
        qty_total = pos.get("qty", 0)
        avg = pos.get("avg_price", 0.0)
        if qty_total <= 0:
            continue
        ltp = ltps.get(tk, avg)
        lots = reconstruct_lots(tk, snapshots, breakup_snaps, all_sells, all_buys, qty_total, avg)
        if not lots:
            # Couldn't reconstruct at all (no xlsx data + no manual fallback)
            rows.append(StockCagr(
                ticker=tk,
                lot_id=f"{tk}#?",
                qty=qty_total,
                avg_price=avg,
                buy_date=None,
                years=None,
                ltp=ltp,
                invested=round(qty_total * avg, 2),
                cur_value=round(qty_total * ltp, 2),
                pnl_pct=round((ltp / avg - 1.0) * 100.0, 2) if avg else 0.0,
                stock_cagr_pct=None,
                nifty_cagr_pct=None,
                alpha_pct=None,
                buy_date_source="missing",
                unreliable_buy_date=True,
            ))
            continue
        for idx, lot in enumerate(lots, start=1):
            years = max((today - lot.buy_date).days / 365.25, 0.001)
            s_cagr = cagr(ltp, lot.buy_price, years) if lot.buy_price > 0 else None
            n_then = nifty_close_on(lot.buy_date)
            n_cagr = cagr(nifty_now, n_then, years) if (n_then and years) else None
            alpha = (s_cagr - n_cagr) if (s_cagr is not None and n_cagr is not None) else None
            unreliable = bool(years * 365.25 < MIN_DAYS_HELD)
            rows.append(StockCagr(
                ticker=tk,
                lot_id=f"{tk}#{idx}",
                qty=lot.qty,
                avg_price=lot.buy_price,
                buy_date=lot.buy_date.isoformat(),
                years=round(years, 3),
                ltp=ltp,
                invested=round(lot.qty * lot.buy_price, 2),
                cur_value=round(lot.qty * ltp, 2),
                pnl_pct=round((ltp / lot.buy_price - 1.0) * 100.0, 2) if lot.buy_price else 0.0,
                stock_cagr_pct=round(s_cagr, 2) if s_cagr is not None else None,
                nifty_cagr_pct=round(n_cagr, 2) if n_cagr is not None else None,
                alpha_pct=round(alpha, 2) if alpha is not None else None,
                buy_date_source=lot.source,
                unreliable_buy_date=unreliable,
            ))

    # Sort: by ticker (alphabetical), then by buy_date (oldest first)
    rows.sort(key=lambda r: (r.ticker, r.buy_date or "9999-99-99"))

    # --- Per-ticker summary (one row per ticker, aggregated across lots) ---
    ticker_summaries: list[StockCagr] = []
    by_ticker: dict[str, list[StockCagr]] = {}
    for r in rows:
        by_ticker.setdefault(r.ticker, []).append(r)
    for tk, lots in by_ticker.items():
        total_qty = sum(l.qty for l in lots)
        total_inv = sum(l.invested for l in lots)
        total_cv = sum(l.cur_value for l in lots)
        # Per-ticker CAGR: simple CAGR on aggregate (single avg, weighted years)
        if total_qty == 0 or total_inv == 0:
            continue
        # Cumulative avg price (truth.json)
        # Per-ticker: find the truth avg for the ticker
        truth_pos = open_positions.get(tk, {})
        truth_avg = truth_pos.get("avg_price", 0.0)
        if truth_avg <= 0 or total_cv <= 0:
            continue
        # Weighted years across this ticker's lots
        valid_lots = [l for l in lots if l.years is not None and l.invested > 0]
        if not valid_lots:
            continue
        wt_years = sum(l.years * l.invested for l in valid_lots) / sum(l.invested for l in valid_lots)
        # Per-ticker simple CAGR
        ticker_return = (total_cv / total_inv) - 1.0
        s_cagr_t = cagr(1.0 + ticker_return, 1.0, wt_years)
        # Per-ticker nifty CAGR (weighted average of per-lot nifty CAGRs)
        wt_nifty = sum((l.nifty_cagr_pct or 0) * l.invested for l in valid_lots) / sum(l.invested for l in valid_lots)
        alpha_t = (s_cagr_t - wt_nifty) if s_cagr_t is not None else None
        # Nifty50 from this ticker's first lot's buy date
        nifty_from_first_lot_pct = None
        ticker_first_buy = min((l.buy_date for l in valid_lots if l.buy_date), default=None)
        if ticker_first_buy and nifty_now:
            t_first = date.fromisoformat(ticker_first_buy)
            t_first_nifty = nifty_close_on(t_first)
            if t_first_nifty and t_first_nifty > 0:
                t_years = max((date.today() - t_first).days / 365.25, 0.001)
                nifty_from_first_lot_pct = cagr(nifty_now, t_first_nifty, t_years)
        unreliable_t = any(l.unreliable_buy_date for l in lots)
        ticker_summaries.append(StockCagr(
            ticker=tk,
            lot_id=f"{tk}#SUMMARY",
            qty=total_qty,
            avg_price=truth_avg,
            buy_date=min((l.buy_date for l in lots if l.buy_date), default=None),
            years=round(wt_years, 3),
            ltp=lots[0].ltp,
            invested=round(total_inv, 2),
            cur_value=round(total_cv, 2),
            pnl_pct=round(ticker_return * 100, 2),
            stock_cagr_pct=round(s_cagr_t, 2) if s_cagr_t is not None else None,
            nifty_cagr_pct=round(wt_nifty, 2) if wt_nifty is not None else None,
            alpha_pct=round(alpha_t, 2) if alpha_t is not None else None,
            buy_date_source="lot_aggregate",
            unreliable_buy_date=unreliable_t,
            nifty_cagr_from_first_buy_pct=(
                round(nifty_from_first_lot_pct, 2)
                if nifty_from_first_lot_pct is not None else None
            ),
            first_buy_date=ticker_first_buy,
        ))

    # --- Portfolio aggregate (across all reliable lots) ---
    valid = [r for r in rows
             if r.years is not None and r.invested > 0
             and not r.unreliable_buy_date]
    if valid and nifty_now:
        total_invested = sum(r.invested for r in valid)
        weighted_years = sum(r.years * r.invested for r in valid) / total_invested
        agg_return = sum(r.cur_value for r in valid) / total_invested - 1.0
        agg_cagr = cagr(1.0 + agg_return, 1.0, weighted_years)
        nifty_cagr_agg = sum(
            (r.nifty_cagr_pct or 0) * r.invested for r in valid
        ) / total_invested
    else:
        weighted_years = None
        agg_cagr = None
        nifty_cagr_agg = None

    # First-investment-date benchmark: Nifty50 CAGR from the earliest lot's
    # buy date to today. This is the apples-to-apples comparison: both
    # started at the same time, with the same money amount in each.
    first_buy_date = None
    nifty_then_first_buy = None
    nifty_cagr_from_first_buy = None
    if valid and nifty_now:
        # Find the earliest buy date across all reliable lots
        first_buy_date = min(
            (r.buy_date for r in valid if r.buy_date),
            default=None,
        )
        if first_buy_date:
            first_buy_date_obj = date.fromisoformat(first_buy_date)
            nifty_then_first_buy = nifty_close_on(first_buy_date_obj)
            if nifty_then_first_buy and nifty_then_first_buy > 0:
                years_from_first_buy = max(
                    (date.today() - first_buy_date_obj).days / 365.25, 0.001
                )
                nifty_cagr_from_first_buy = cagr(
                    nifty_now, nifty_then_first_buy, years_from_first_buy
                )

    # Also: what if you'd invested the whole book in Nifty50 on the
    # first-buy date? This is the "if you had just bought index" counterfactual.
    # The Nifty50 CAGR over the same period *is* the counterfactual return
    # for a single lump-sum investment. (Not a per-lot DCA simulation —
    # that's XIRR, which we'd need buy dates for each cashflow.)
    counterfactual_nifty_value = None
    if total_invested and nifty_then_first_buy and nifty_now:
        counterfactual_nifty_value = total_invested * (nifty_now / nifty_then_first_buy)

    missing = [r.ticker for r in rows if r.buy_date is None and r.lot_id.endswith("#?")]
    unreliable = sorted({r.ticker for r in rows if r.unreliable_buy_date})

    # Totals across ALL open positions (reliable + unreliable) — matches
    # what the broker shows as "current value". Used for the headline
    # numbers, separate from the CAGR-restricted aggregate above.
    all_open_rows = [r for r in rows if r.years is not None and r.invested > 0]
    all_invested = sum(r.invested for r in all_open_rows)
    all_cur_value = sum(r.cur_value for r in all_open_rows)
    all_pnl = all_cur_value - all_invested
    all_pnl_pct = (all_pnl / all_invested * 100) if all_invested else 0.0

    return {
        "asof": today.isoformat(),
        "ltp_source": ltp_source,
        "nifty_now": nifty_now,
        "nifty_history_points": len(_nifty_history()),
        "nifty_first_date": _nifty_history()[0][0].isoformat() if _nifty_history() else None,
        "nifty_last_date": _nifty_history()[-1][0].isoformat() if _nifty_history() else None,
        "lots": [asdict(r) for r in rows],
        "ticker_summaries": [asdict(r) for r in ticker_summaries],
        "aggregate": {
            "weighted_years": round(weighted_years, 3) if weighted_years else None,
            "stock_cagr_pct": round(agg_cagr, 2) if agg_cagr is not None else None,
            # Per-lot weighted Nifty50 CAGR (each lot's window)
            "nifty_cagr_pct": round(nifty_cagr_agg, 2) if nifty_cagr_agg is not None else None,
            "alpha_pct": round((agg_cagr - nifty_cagr_agg), 2)
                         if (agg_cagr is not None and nifty_cagr_agg is not None) else None,
            # First-investment-date benchmark (the apples-to-apples comparison)
            "first_buy_date": first_buy_date,
            "nifty_then_first_buy": nifty_then_first_buy,
            "nifty_cagr_from_first_buy_pct": (
                round(nifty_cagr_from_first_buy, 2)
                if nifty_cagr_from_first_buy is not None else None
            ),
            "alpha_from_first_buy_pct": (
                round((agg_cagr - nifty_cagr_from_first_buy), 2)
                if (agg_cagr is not None and nifty_cagr_from_first_buy is not None) else None
            ),
            # Counterfactual: what if you'd lump-sum-invested the whole
            # book in Nifty50 on the first buy date?
            "counterfactual_nifty_value": (
                round(counterfactual_nifty_value, 2)
                if counterfactual_nifty_value is not None else None
            ),
            "counterfactual_diff": (
                round(sum(r.cur_value for r in valid) - counterfactual_nifty_value, 2)
                if (counterfactual_nifty_value is not None and valid) else None
            ),
            "total_invested": round(sum(r.invested for r in valid), 2),
            "total_cur_value": round(sum(r.cur_value for r in valid), 2),
            # Totals across ALL open lots (matches broker view)
            "all_total_invested": round(all_invested, 2),
            "all_total_cur_value": round(all_cur_value, 2),
            "all_total_pnl": round(all_pnl, 2),
            "all_total_pnl_pct": round(all_pnl_pct, 2),
            "missing_buy_dates": missing,
            "unreliable_buy_dates": unreliable,
        },
    }

# ---------------------------------------------------------------------------
# Cache (mtime-based, mirrors webapp/cache.py pattern)
# ---------------------------------------------------------------------------

def _cache_key(tm: float, ntm: float) -> str:
    return f"{tm:.0f}_{ntm:.0f}"


def _read_cache(key: str) -> Optional[dict]:
    if not CACHE_FILE.exists():
        return None
    try:
        with CACHE_FILE.open() as fh:
            d = json.load(fh)
        return d.get(key)
    except Exception:
        return None


def _write_cache(key: str, payload: dict) -> None:
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    cur: dict = {}
    if CACHE_FILE.exists():
        try:
            with CACHE_FILE.open() as fh:
                cur = json.load(fh)
        except Exception:
            cur = {}
    cur[key] = payload
    # Keep only last 20 entries
    if len(cur) > 20:
        for k in list(cur.keys())[:-20]:
            del cur[k]
    with CACHE_FILE.open("w") as fh:
        json.dump(cur, fh, indent=2, default=str)


def get_equity_cagr(force: bool = False) -> dict:
    """Public entry point. Returns the full CAGR report. Cached by
    (truth_mtime, nifty_csv_mtime) so it re-runs when inputs change."""
    tm = truth_mtime()
    ntm = nifty_mtime()
    key = _cache_key(tm, ntm)
    if not force:
        cached = _read_cache(key)
        if cached is not None:
            log.info("cagr cache hit (key=%s)", key)
            return cached
    log.info("cagr cache miss (key=%s) — computing", key)
    result = compute_equity_cagr()
    _write_cache(key, result)
    return result


# ---------------------------------------------------------------------------
# Pretty-print
# ---------------------------------------------------------------------------

def print_report(report: Optional[dict] = None) -> None:
    if report is None:
        report = get_equity_cagr()
    print()
    print("=" * 120)
    print(f"  EQUITY CAGR vs NIFTY50  ·  as of {report['asof']}  ·  LTP source: {report['ltp_source']}")
    print("=" * 120)
    print(f"  Nifty50 reference close: {report['nifty_now']:.2f}  ({report['nifty_last_date']})")
    print(f"  Nifty50 history: {report['nifty_history_points']} pts  "
          f"({report['nifty_first_date']} → {report['nifty_last_date']})")
    print("-" * 120)
    # Per-ticker summary first
    print("  -- Per-ticker summary (one row per ticker, lots aggregated) --")
    hdr = f"  {'TICKER':<14}{'QTY':>6}{'AVG':>10}{'LTP':>10}{'BUY DATE':>12}{'YRS':>8}{'CAGR%':>10}{'NIFTY%':>10}{'α%':>10}"
    print(hdr)
    print("-" * 120)
    for s in report.get("ticker_summaries", []):
        def _fmt_pct(v):
            if v is None: return "    --  "
            if v > 9999:  return "  >9999"
            if v < -9999: return "  <-9999"
            return f"{v:>+8.2f}"
        cagr_s = _fmt_pct(s["stock_cagr_pct"])
        nift_s = _fmt_pct(s["nifty_cagr_pct"])
        alph_s = _fmt_pct(s["alpha_pct"])
        bd = s["buy_date"] or "MISSING"
        yrs = f"{s['years']:.3f}" if s["years"] is not None else "   --  "
        print(f"  {s['ticker']:<14}{s['qty']:>6}{s['avg_price']:>10.2f}{s['ltp']:>10.2f}"
              f"{bd:>12}{yrs:>8}{cagr_s:>10}{nift_s:>10}{alph_s:>10}")
    print("-" * 120)
    # Per-lot detail
    print()
    print("  -- Per-LOT detail (one row per buy lot, reconstructed from xlsx) --")
    hdr = f"  {'LOT':<20}{'QTY':>6}{'BUY PRICE':>11}{'LTP':>10}{'BUY DATE':>12}{'YRS':>8}{'CAGR%':>10}{'NIFTY%':>10}{'α%':>10}  {'SOURCE':<18}"
    print(hdr)
    print("-" * 120)
    for s in report.get("lots", []):
        def _fmt_pct(v):
            if v is None: return "    --  "
            if v > 9999:  return "  >9999"
            if v < -9999: return "  <-9999"
            return f"{v:>+8.2f}"
        cagr_s = _fmt_pct(s["stock_cagr_pct"])
        nift_s = _fmt_pct(s["nifty_cagr_pct"])
        alph_s = _fmt_pct(s["alpha_pct"])
        bd = s["buy_date"] or "MISSING"
        yrs = f"{s['years']:.3f}" if s["years"] is not None else "   --  "
        src = s["buy_date_source"]
        print(f"  {s['lot_id']:<20}{s['qty']:>6}{s['avg_price']:>11.2f}{s['ltp']:>10.2f}"
              f"{bd:>12}{yrs:>8}{cagr_s:>10}{nift_s:>10}{alph_s:>10}  {src:<18}")
    print("-" * 120)
    a = report["aggregate"]
    def _af(v):
        if v is None: return 0.0
        if v > 9999:  return 9999.0
        if v < -9999: return -9999.0
        return v
    print(f"  {'AGGREGATE':<20}{'':>6}{'':>11}{'':>10}{'':>12}"
          f"{a['weighted_years'] or 0:>8.3f}"
          f"{_af(a['stock_cagr_pct']):>+10.2f}"
          f"{_af(a['nifty_cagr_pct']):>+10.2f}"
          f"{_af(a['alpha_pct']):>+10.2f}")
    print("=" * 120)
    print(f"  Invested: ₹{a['total_invested']:,.0f}   "
          f"Current: ₹{a['total_cur_value']:,.0f}   "
          f"Total P&L: ₹{a['total_cur_value']-a['total_invested']:+,.0f}")
    if a["unreliable_buy_dates"]:
        print(f"  ⚠ Lots with <{MIN_DAYS_HELD} days held (excluded from aggregate): "
              f"{', '.join(a['unreliable_buy_dates'])}")
    if a["missing_buy_dates"]:
        print(f"  ⚠ Missing buy dates: {', '.join(a['missing_buy_dates'])}")
    print("=" * 120)
    print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Equity CAGR vs Nifty50")
    p.add_argument("--json", action="store_true", help="print JSON instead of table")
    p.add_argument("--force", action="store_true", help="bypass cache")
    args = p.parse_args()
    rep = get_equity_cagr(force=args.force)
    if args.json:
        print(json.dumps(rep, indent=2, default=str))
    else:
        print_report(rep)
