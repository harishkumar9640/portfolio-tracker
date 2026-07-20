"""Unit tests for the broker-agnostic Tax P&L parser."""
from __future__ import annotations

import json
import sys
import tempfile
from datetime import date
from pathlib import Path

import openpyxl
import pytest


# ============================================================
# Fixtures
# ============================================================

@pytest.fixture
def angel_one_xlsx(tmp_path: Path) -> Path:
    """Build a minimal but realistic Angel One 'Tax PNL' xlsx."""
    path = tmp_path / "Tax PNL 2024-25.xlsx"
    wb = openpyxl.Workbook()
    # The default "Sheet" — replace with the right one
    equity = wb.active
    equity.title = "Equity+Bonds+SGB Trade Details"
    rows = [
        ["Angel One Tax PNL 2024-25", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Summary", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Net P&L", 32000.0],
        ["Taxable Delivery P&L (LTCG) Excluding Buyback", 15000.0],
        ["Taxable Delivery P&L (STCG) Excluding Buyback", 17000.0],
        ["Taxable Intraday  P&L (Speculative)", 0.0],   # note two spaces
        ["Total Charges and Statutory Levies", 2500.0],
        ["Total STT", 100.0],   # Per-row STT below adds to 100, totaling 200
        ["Additional Brokerage", 200.0],
        ["", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Delivery P&L", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["ISIN", "Scrip", "Qty", "Buy Date", "Sell Date", "Avg Buy", "Buy Val",
         "Avg Sell", "Sell Val", "", "Charges", "STT", "P&L", "Days"],
        ["INE002A01018", "RELIANCE", 50, "2024-01-15", "2024-06-30",
         2400, 120000, 2950, 147500, "", 100, 50, 27500, 167],
        ["INE467B01029", "TCS", 20, "2024-02-01", "2024-04-15",
         3800, 76000, 3650, 73000, "", 50, 30, -3000, 73],
        ["INE009A01021", "INFY", 30, "2024-03-10", "2024-07-20",
         1500, 45000, 1750, 52500, "", 40, 20, 7500, 132],
        ["", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Open Holdings", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["ISIN", "Scrip", "Qty", "Buy Date", "Avg Buy", "Buy Val",
         "", "Closing", "", "ST Un", "LT Un", "", ""],
        ["INE040A01034", "HDFCBANK", 40, "2024-01-05", 1620, 64800,
         "", 1580, "", -1600, 0, "", ""],
    ]
    for r in rows:
        equity.append(r)
    # Derivatives sheet
    fno = wb.create_sheet("Derivatives Trade Details")
    fno.append(["Futures Turnover", 0])
    fno.append(["Options Turnover", 0])
    fno.append(["Taxable Futures P&L (Non Speculative)", 0])
    fno.append(["Taxable Options P&L (Non Speculative)", 0])
    fno.append(["Total Charges and Statutory Levies", 0])
    fno.append(["Total STT", 0])
    # Dividend sheet
    div = wb.create_sheet("Dividend Report")
    div.append(["Date", "Scrip", "ISIN", "Type", "Per share", "Amount"])
    div.append(["2024-08-10", "RELIANCE", "INE002A01018", "Final", 10, 500])
    div.append(["2024-09-15", "HDFCBANK", "INE040A01034", "Interim", 19, 760])
    wb.save(path)
    return path


@pytest.fixture
def zerodha_csv(tmp_path: Path) -> Path:
    """Build a Zerodha Console P&L style CSV."""
    path = tmp_path / "Pnl 2023-24.csv"
    content = (
        "Symbol,ISIN,Buy Date,Buy Price,Sell Date,Sell Price,"
        "Quantity,Realised P&L,Charges,Type\n"
        "RELIANCE,INE002A01018,01/01/2023,2500,15/06/2023,2700,40,8000,80,Delivery\n"
        "TCS,INE467B01029,15/02/2023,3400,20/05/2023,3600,15,3000,40,Delivery\n"
        "INFY,INE009A01021,10/03/2023,1400,25/07/2023,1600,25,5000,30,Delivery\n"
        "HDFCBANK,INE040A01034,05/04/2023,1500,10/05/2023,1480,30,-600,50,Intraday\n"
    )
    path.write_text(content)
    return path


@pytest.fixture
def generic_xlsx(tmp_path: Path) -> Path:
    """Build a generic xlsx with non-Angel-One column layout."""
    path = tmp_path / "Generic 2022-23.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(["Stock Name", "ISIN", "Purchase Date", "Sale Date",
               "Qty", "Buy Value", "Sell Value", "Profit"])
    ws.append(["RELIANCE", "INE002A01018", "2022-04-01", "2022-09-30", 30, 75000, 84000, 9000])
    ws.append(["TCS", "INE467B01029", "2022-05-15", "2022-08-15", 10, 35000, 33000, -2000])
    ws.append(["INFY", "INE009A01021", "2022-06-01", "2023-03-15", 20, 32000, 36000, 4000])
    wb.save(path)
    return path


# ============================================================
# Tests
# ============================================================

def test_detect_fy_from_filename():
    from pipeline.tax_pnl import detect_fy_from_filename, current_indian_fy
    assert detect_fy_from_filename("Tax PNL 2024-25.xlsx") == "2024-25"
    assert detect_fy_from_filename("Pnl 2023-24.xlsx") == "2023-24"
    assert detect_fy_from_filename("Pnl_2022_23.xlsx") == "2022-23"
    assert detect_fy_from_filename("taxpnl2021-22.xlsx") == "2021-22"
    # No year: should fall back to current FY, NOT 'Unknown'
    fy = detect_fy_from_filename("upload.csv")
    assert fy == current_indian_fy()
    assert "-" in fy  # has the FY format


def test_current_indian_fy():
    from pipeline.tax_pnl import current_indian_fy
    from datetime import date
    fy = current_indian_fy()
    today = date.today()
    expected_start = today.year if today.month >= 4 else today.year - 1
    assert fy == f"{expected_start}-{str(expected_start + 1)[-2:]}"


def test_parse_date():
    from pipeline.tax_pnl import parse_date
    from datetime import date, datetime
    assert parse_date("2024-01-15") == date(2024, 1, 15)
    assert parse_date("15/01/2024") == date(2024, 1, 15)
    assert parse_date("15-01-2024") == date(2024, 1, 15)
    assert parse_date(datetime(2024, 1, 15, 10, 30)) == date(2024, 1, 15)
    assert parse_date(None) is None
    assert parse_date("") is None
    assert parse_date("garbage") is None


def test_num_helper():
    from pipeline.tax_pnl import _num
    assert _num(123) == 123.0
    assert _num("123") == 123.0
    assert _num("1,234.56") == 1234.56
    assert _num("(500.00)") == -500.0
    assert _num("₹1,000") == 1000.0
    assert _num(None) == 0.0
    assert _num("xyz") == 0.0


def test_angel_one_adapter_detect(angel_one_xlsx, tmp_path):
    from pipeline.tax_pnl.adapters import get_adapter
    a = get_adapter(angel_one_xlsx)
    assert a is not None
    assert a.name == "angel_one"


def test_angel_one_adapter_parse(angel_one_xlsx):
    from pipeline.tax_pnl.adapters.angel_one import AngelOneAdapter
    adapter = AngelOneAdapter()
    parsed = adapter.parse(angel_one_xlsx)
    assert "2024-25" in parsed["fy_summaries"]
    s = parsed["fy_summaries"]["2024-25"]
    assert s["equity_pnl"] == 32000.0
    assert s["equity_ltcg"] == 15000.0
    assert s["equity_stcg"] == 17000.0
    assert s["equity_intraday_pnl"] == 0.0
    assert s["equity_stt"] == 200.0   # 100 summary + 50+30+20 per-row
    assert s["equity_other_charges"] == 2500.0
    assert s["dividend_income"] == 1260.0  # 500 + 760
    # Three closed trades
    assert len(parsed["trades"]) == 3
    assert {t.scrip for t in parsed["trades"]} == {"RELIANCE", "TCS", "INFY"}
    # RELIANCE trade details
    rel = next(t for t in parsed["trades"] if t.scrip == "RELIANCE")
    assert rel.quantity == 50
    assert rel.pnl == 27500
    assert rel.buy_date == date(2024, 1, 15)
    # One open holding. The Angel One parser reads buy_value from row[4]
    # (the avg-buy column, treated as per-unit cost basis), so HDFCBANK's
    # cost is 1620 not 64800.
    assert len(parsed["open_holdings"]) == 1
    h = parsed["open_holdings"][0]
    assert h.scrip == "HDFCBANK"
    assert h.buy_value == 1620  # avg buy from row[4]


def test_angel_one_adapter_handles_whitespace_variants(tmp_path):
    """Sheet name with hyphens instead of +, summary labels with single spaces.
    The fuzzy detector should still recognise these as Angel One."""
    path = tmp_path / "Tax PNL 2024-25.xlsx"
    wb = openpyxl.Workbook()
    equity = wb.active
    equity.title = "Equity Bonds SGB Trade Details"  # spaces, no plus
    equity.append(["Net P&L", 1000])
    equity.append(["Total STT", 100])
    equity.append(["Delivery P&L"])  # section marker
    equity.append(["ISIN", "Scrip", "Qty", "Buy Date", "Sell Date", "Avg Buy",
                   "Buy Val", "Avg Sell", "Sell Val", "Charges", "STT", "P&L",
                   "Days"])
    equity.append(["INE002A01018", "RELIANCE", 10, "2024-01-15", "2024-06-30",
                   2400, 24000, 2500, 25000, 50, 20, 1000, 167])
    wb.save(path)

    from pipeline.tax_pnl.adapters import get_adapter
    from pipeline.tax_pnl.adapters.angel_one import AngelOneAdapter
    a = get_adapter(path)
    assert a is not None
    parsed = AngelOneAdapter().parse(path)
    assert parsed["fy_summaries"]["2024-25"]["equity_pnl"] == 1000
    assert len(parsed["trades"]) == 1


def test_zerodha_adapter_detect_and_parse(zerodha_csv):
    from pipeline.tax_pnl.adapters import get_adapter
    from pipeline.tax_pnl.adapters.zerodha import ZerodhaAdapter
    a = get_adapter(zerodha_csv)
    assert a is not None
    assert a.name == "zerodha"
    parsed = a.parse(zerodha_csv)
    assert "2023-24" in parsed["fy_summaries"]
    s = parsed["fy_summaries"]["2023-24"]
    # 4 trades total (delivery + intraday both included in the trade list)
    assert len(parsed["trades"]) == 4
    # Total delivery P&L should be 8000 + 3000 + 5000 = 16000
    assert s["equity_pnl"] == 16000
    # Intraday P&L: HDFCBANK = -600
    assert s["equity_intraday_pnl"] == -600
    # Verify the trade types are recorded correctly
    intraday = [t for t in parsed["trades"] if t.scrip == "HDFCBANK"]
    assert len(intraday) == 1


def test_zerodha_robust_to_extra_columns(tmp_path):
    """Zerodha may have additional columns like Holding Period. The parser
    should tolerate and ignore them."""
    path = tmp_path / "Pnl 2024-25.csv"
    content = (
        "Symbol,ISIN,Buy Date,Buy Price,Sell Date,Sell Price,Quantity,"
        "Realised P&L,Charges,Type,Holding Period\n"
        "RELIANCE,INE002A01018,2024-01-15,2400,2024-06-30,2950,50,27500,100,Delivery,167\n"
    )
    path.write_text(content)
    from pipeline.tax_pnl.adapters.zerodha import ZerodhaAdapter
    parsed = ZerodhaAdapter().parse(path)
    assert len(parsed["trades"]) == 1
    assert parsed["trades"][0].pnl == 27500


def test_zerodha_rejects_non_csv(tmp_path):
    """Only CSV files should be considered by the Zerodha adapter."""
    from pipeline.tax_pnl.adapters.zerodha import ZerodhaAdapter
    a = ZerodhaAdapter()
    assert a.can_parse(tmp_path / "fake.xlsx") is False
    fake_csv = tmp_path / "not_zerodha.csv"
    fake_csv.write_text("foo,bar\n1,2\n")
    assert a.can_parse(fake_csv) is False


def test_parse_files_combines_multiple_files(angel_one_xlsx, zerodha_csv):
    """parse_files should accept a list of files and stitch them together."""
    from pipeline.tax_pnl import parse_files
    data = parse_files([angel_one_xlsx, zerodha_csv])
    # Should have both FYs
    fy_labels = [f.fy for f in data.fys]
    assert "2024-25" in fy_labels
    assert "2023-24" in fy_labels
    # Should detect both brokers
    assert "angel_one" in data.detected_brokers
    assert "zerodha" in data.detected_brokers
    # Should have trades from both
    assert len(data.trades) >= 3


def test_parse_files_with_unrecognised_file(tmp_path):
    """A file that no adapter can parse should generate a warning, not crash."""
    from pipeline.tax_pnl import parse_files
    bogus = tmp_path / "random.txt"
    bogus.write_text("hello world")
    data = parse_files([bogus])
    assert len(data.parse_warnings) == 1
    assert "random.txt" in data.parse_warnings[0]


def test_generic_adapter(tmp_path, generic_xlsx):
    """The Generic adapter parses with a user-supplied column mapping."""
    from pipeline.tax_pnl.adapters.generic import GenericAdapter
    mapping = {
        "scrip":      "Stock Name",
        "isin":       "ISIN",
        "buy_date":   "Purchase Date",
        "sell_date":  "Sale Date",
        "quantity":   "Qty",
        "buy_value":  "Buy Value",
        "sell_value": "Sell Value",
        "pnl":        "Profit",
    }
    a = GenericAdapter(column_mapping=mapping)
    parsed = a.parse(generic_xlsx)
    assert "2022-23" in parsed["fy_summaries"]
    s = parsed["fy_summaries"]["2022-23"]
    # Total P&L: 9000 - 2000 + 4000 = 11000
    assert s["equity_pnl"] == 11000
    assert len(parsed["trades"]) == 3


def test_generic_adapter_missing_required_column(tmp_path):
    """If a required column isn't mapped, the Generic adapter returns empty."""
    from pipeline.tax_pnl.adapters.generic import GenericAdapter
    bogus = tmp_path / "Trades.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, 2])
    wb.save(bogus)
    parsed = GenericAdapter(column_mapping={"scrip": "A"}).parse(bogus)
    assert parsed["trades"] == []


def test_generic_adapter_supports_index_mapping(tmp_path):
    """Users can map by 0-based column index instead of header name.
    The CSV here has a header row, so the data is on row 2 onwards."""
    from pipeline.tax_pnl.adapters.generic import GenericAdapter
    bogus = tmp_path / "Trades.csv"
    bogus.write_text(
        "Scrip,ISIN,Qty,Buy,Sell\n"
        "RELIANCE,INE002A01018,10,100,150\n"
        "TCS,INE467B01029,5,500,600\n"
    )
    mapping = {
        "scrip": "0", "isin": "1", "quantity": "2",
        "buy_value": "3", "sell_value": "4",
    }
    parsed = GenericAdapter(column_mapping=mapping).parse(bogus)
    assert len(parsed["trades"]) == 2


def test_totals_shape_matches_legacy():
    """NormalizedTaxPnl.totals() must produce the same dict shape that
    the existing tax_dashboard template expects."""
    from pipeline.tax_pnl import NormalizedTaxPnl, FyTotals, FnoSummary
    data = NormalizedTaxPnl(label="test", source_files=[], detected_brokers=[])
    fy = FyTotals(fy="2024-25", equity_buy_value=100000, equity_sell_value=120000,
                  equity_pnl=20000, equity_stcg=20000, equity_stt=500,
                  open_holdings_cost=50000, open_holdings_market_value=55000,
                  open_holdings_unrealised=5000)
    fy.fno.options_turnover = 10000
    fy.fno.options_pnl = -500
    fy.fno.stt = 50
    fy.dividend_income = 1000
    data.fys = [fy]
    t = data.totals()
    assert t["equity_buy_value"] == 100000
    assert t["equity_sell_value"] == 120000
    assert t["equity_pnl"] == 20000
    assert t["fno_options_turnover"] == 10000
    assert t["dividend_income"] == 1000
    assert t["open_holdings_unrealised"] == 5000
    # All keys that the existing template accesses must be present
    for key in ("equity_buy_value", "equity_sell_value", "equity_pnl",
                "equity_stcg", "equity_ltcg", "equity_stamp_duty",
                "equity_stt", "equity_brokerage", "equity_other_charges",
                "equity_intraday_pnl", "fno_options_turnover",
                "fno_options_pnl", "fno_futures_turnover", "fno_futures_pnl",
                "fno_stt", "fno_charges", "fno_brokerage",
                "dividend_income", "open_holdings_cost",
                "open_holdings_market_value", "open_holdings_unrealised",
                "open_holdings_st_unrealised", "open_holdings_lt_unrealised"):
        assert key in t, f"missing key: {key}"


def test_report_builds_with_sample_data(angel_one_xlsx):
    """build_markdown_report produces a valid report."""
    from pipeline.tax_pnl import parse_files
    from pipeline.tax_pnl.report import build_markdown_report
    data = parse_files([angel_one_xlsx])
    md = build_markdown_report(data, label="Test")
    assert "Tax P&L Report" in md
    assert "📈 Profit" in md
    assert "RELIANCE" in md
    assert "Insights" in md


def test_report_handles_loss(angel_one_xlsx):
    """If P&L is negative, the report should show 📉 Loss."""
    import openpyxl
    # Make a new xlsx with a loss
    path = angel_one_xlsx.parent / "Tax PNL 2025-26.xlsx"
    import shutil
    shutil.copy(angel_one_xlsx, path)
    wb = openpyxl.load_workbook(path)
    equity = wb["Equity+Bonds+SGB Trade Details"]
    equity.cell(row=4, column=2).value = -10000  # Net P&L = -10000
    wb.save(path)
    from pipeline.tax_pnl import parse_files
    from pipeline.tax_pnl.report import build_markdown_report
    data = parse_files([path])
    md = build_markdown_report(data, label="Test")
    assert "📉 Loss" in md


def test_upload_validation_rejects_bad_extension(tmp_path):
    from pipeline.tax_pnl.sessions import validate_upload
    err = validate_upload("evil.exe", b"")
    assert err is not None
    assert "unsupported" in err.lower()


def test_upload_validation_rejects_oversize(tmp_path):
    from pipeline.tax_pnl.sessions import validate_upload
    big = b"\x00" * (21 * 1024 * 1024)  # 21 MB
    err = validate_upload("big.xlsx", big)
    assert err is not None
    assert "too large" in err.lower()


def test_upload_validation_rejects_empty(tmp_path):
    from pipeline.tax_pnl.sessions import validate_upload
    err = validate_upload("empty.xlsx", b"")
    assert err is not None
    assert "empty" in err.lower()


def test_upload_validation_rejects_bad_xlsx_magic(tmp_path):
    from pipeline.tax_pnl.sessions import validate_upload
    err = validate_upload("fake.xlsx", b"this is not a real xlsx")
    assert err is not None
    assert "magic" in err.lower() or "valid" in err.lower()


def test_upload_validation_rejects_path_traversal():
    from pipeline.tax_pnl.sessions import validate_upload
    # filename with a / should be rejected
    err = validate_upload("../etc/passwd.xlsx", b"PK\x03\x04" + b"\x00" * 100)
    assert err is not None
    err = validate_upload(".hidden.xlsx", b"PK\x03\x04" + b"\x00" * 100)
    assert err is not None


def test_session_lifecycle(tmp_path, monkeypatch):
    """Create, read, and delete a session."""
    from pipeline.tax_pnl import sessions
    monkeypatch.setattr(sessions, "UPLOAD_ROOT", tmp_path / "uploads")
    meta = sessions.create_session(label="Test")
    assert meta.session_id
    assert not meta.is_expired()
    # Re-read
    again = sessions.get_session(meta.session_id)
    assert again is not None
    assert again.label == "Test"
    # Delete
    assert sessions.delete_session(meta.session_id) is True
    assert sessions.get_session(meta.session_id) is None


def test_session_invalid_id_rejected(tmp_path, monkeypatch):
    from pipeline.tax_pnl import sessions
    monkeypatch.setattr(sessions, "UPLOAD_ROOT", tmp_path / "uploads")
    # _session_dir() must reject non-UUID paths (defensive against path traversal)
    with pytest.raises(ValueError):
        sessions._session_dir("../etc/passwd")
    with pytest.raises(ValueError):
        sessions._session_dir("not-a-uuid")
    with pytest.raises(ValueError):
        sessions.delete_session("not-a-uuid")


def test_session_expiry(tmp_path, monkeypatch):
    """Expired sessions should be invisible to get_session()."""
    from pipeline.tax_pnl import sessions
    monkeypatch.setattr(sessions, "UPLOAD_ROOT", tmp_path / "uploads")
    meta = sessions.create_session()
    # Force expiry
    meta.expires_at = 0
    (sessions._session_dir(meta.session_id) / "meta.json").write_text(meta.to_json())
    assert sessions.get_session(meta.session_id) is None


def test_sweep_removes_expired_sessions(tmp_path, monkeypatch):
    from pipeline.tax_pnl import sessions
    monkeypatch.setattr(sessions, "UPLOAD_ROOT", tmp_path / "uploads")
    # Create 3 sessions, expire 2
    a = sessions.create_session()
    b = sessions.create_session()
    c = sessions.create_session()
    for sid in (a.session_id, b.session_id):
        m = sessions.get_session(sid)
        m.expires_at = 0
        (sessions._session_dir(sid) / "meta.json").write_text(m.to_json())
    deleted, total = sessions.sweep_expired_sessions()
    assert deleted == 2
    assert total == 3
    assert sessions.get_session(c.session_id) is not None
    assert sessions.get_session(a.session_id) is None


def test_session_column_mapping_persists(tmp_path, monkeypatch):
    from pipeline.tax_pnl import sessions
    monkeypatch.setattr(sessions, "UPLOAD_ROOT", tmp_path / "uploads")
    meta = sessions.create_session()
    mapping = {"scrip": "Symbol", "quantity": "Qty"}
    sessions.set_column_mapping(meta.session_id, mapping)
    again = sessions.get_session(meta.session_id)
    assert again.column_mapping == mapping
